"""Tests for the XLSX tool."""

from __future__ import annotations

from unittest.mock import patch

import pytest

from anteroom.tools.office_xlsx import (
    _MAX_EDIT_OPS,
    _MAX_ROWS,
    AVAILABLE,
    DEFINITION,
    _sanitize_cell_value,
    handle,
    set_working_dir,
)

_needs_openpyxl = pytest.mark.skipif(not AVAILABLE, reason="requires openpyxl: pip install anteroom[office]")


@pytest.fixture(autouse=True)
def _set_working_dir(tmp_path):
    set_working_dir(str(tmp_path))
    yield


class TestDefinition:
    def test_name(self):
        assert DEFINITION["name"] == "xlsx"

    def test_required_params(self):
        assert "action" in DEFINITION["parameters"]["required"]
        assert "path" in DEFINITION["parameters"]["required"]


@_needs_openpyxl
class TestCreate:
    @pytest.mark.asyncio
    async def test_create_simple(self, tmp_path):
        result = await handle(
            action="create",
            path="test.xlsx",
            sheets=[{"name": "Data", "headers": ["Name", "Age"], "rows": [["Alice", 30], ["Bob", 25]]}],
        )
        assert "error" not in result
        assert result["sheets_created"] == 1
        assert result["total_rows"] == 3  # 1 header + 2 data
        assert (tmp_path / "test.xlsx").exists()

    @pytest.mark.asyncio
    async def test_create_multiple_sheets(self, tmp_path):
        result = await handle(
            action="create",
            path="multi.xlsx",
            sheets=[
                {"name": "Sheet1", "rows": [["a", "b"]]},
                {"name": "Sheet2", "rows": [["c", "d"]]},
            ],
        )
        assert "error" not in result
        assert result["sheets_created"] == 2

    @pytest.mark.asyncio
    async def test_create_no_sheets(self):
        result = await handle(action="create", path="empty.xlsx")
        assert "error" in result

    @pytest.mark.asyncio
    async def test_create_too_many_rows(self):
        rows = [[i] for i in range(_MAX_ROWS + 1)]
        result = await handle(action="create", path="big.xlsx", sheets=[{"name": "Big", "rows": rows}])
        assert "error" in result
        assert "Too many rows" in result["error"]


@_needs_openpyxl
class TestRead:
    @pytest.mark.asyncio
    async def test_read_simple(self, tmp_path):
        await handle(
            action="create",
            path="read_me.xlsx",
            sheets=[{"name": "Data", "rows": [["Alice", 30], ["Bob", 25]]}],
        )
        result = await handle(action="read", path="read_me.xlsx")
        assert "error" not in result
        assert "Alice" in result["content"]
        assert result["rows_read"] == 2

    @pytest.mark.asyncio
    async def test_read_specific_sheet(self, tmp_path):
        await handle(
            action="create",
            path="multi.xlsx",
            sheets=[
                {"name": "First", "rows": [["a"]]},
                {"name": "Second", "rows": [["b"]]},
            ],
        )
        result = await handle(action="read", path="multi.xlsx", sheet_name="Second")
        assert "error" not in result
        assert "b" in result["content"]

    @pytest.mark.asyncio
    async def test_read_missing_sheet(self, tmp_path):
        await handle(
            action="create",
            path="test.xlsx",
            sheets=[{"name": "Data", "rows": [["x"]]}],
        )
        result = await handle(action="read", path="test.xlsx", sheet_name="NonExistent")
        assert "error" in result
        assert "not found" in result["error"]

    @pytest.mark.asyncio
    async def test_read_not_found(self):
        result = await handle(action="read", path="missing.xlsx")
        assert "error" in result

    @pytest.mark.asyncio
    async def test_read_corrupt_file(self, tmp_path):
        corrupt = tmp_path / "corrupt.xlsx"
        corrupt.write_bytes(b"not an xlsx file")
        result = await handle(action="read", path="corrupt.xlsx")
        assert "error" in result

    @pytest.mark.asyncio
    async def test_read_with_range(self, tmp_path):
        await handle(
            action="create",
            path="range.xlsx",
            sheets=[{"name": "Data", "rows": [["a", "b", "c"], ["1", "2", "3"], ["4", "5", "6"]]}],
        )
        result = await handle(action="read", path="range.xlsx", cell_range="A1:B2")
        assert "error" not in result
        assert result["rows_read"] == 2


@_needs_openpyxl
class TestEdit:
    @pytest.mark.asyncio
    async def test_edit_cells(self, tmp_path):
        await handle(
            action="create",
            path="edit_me.xlsx",
            sheets=[{"name": "Data", "rows": [["old_value"]]}],
        )
        result = await handle(
            action="edit",
            path="edit_me.xlsx",
            updates=[{"cell": "A1", "value": "new_value"}],
        )
        assert "error" not in result
        assert result["cells_updated"] == 1

    @pytest.mark.asyncio
    async def test_edit_append_rows(self, tmp_path):
        await handle(
            action="create",
            path="append.xlsx",
            sheets=[{"name": "Data", "rows": [["a"]]}],
        )
        result = await handle(
            action="edit",
            path="append.xlsx",
            append_rows=[["b"], ["c"]],
        )
        assert "error" not in result
        assert result["rows_appended"] == 2

    @pytest.mark.asyncio
    async def test_edit_add_sheet(self, tmp_path):
        await handle(
            action="create",
            path="sheets.xlsx",
            sheets=[{"name": "Original", "rows": [["x"]]}],
        )
        result = await handle(
            action="edit",
            path="sheets.xlsx",
            add_sheets=[{"name": "NewSheet", "rows": [["y"]]}],
        )
        assert "error" not in result
        assert result["sheets_added"] == 1

    @pytest.mark.asyncio
    async def test_edit_not_found(self):
        result = await handle(
            action="edit",
            path="missing.xlsx",
            updates=[{"cell": "A1", "value": "x"}],
        )
        assert "error" in result

    @pytest.mark.asyncio
    async def test_edit_no_operations(self, tmp_path):
        await handle(
            action="create",
            path="noop.xlsx",
            sheets=[{"name": "Data", "rows": [["x"]]}],
        )
        result = await handle(action="edit", path="noop.xlsx")
        assert "error" in result

    @pytest.mark.asyncio
    async def test_edit_too_many_append_rows(self, tmp_path):
        await handle(
            action="create",
            path="big.xlsx",
            sheets=[{"name": "Data", "rows": [["x"]]}],
        )
        rows = [[i] for i in range(_MAX_ROWS + 1)]
        result = await handle(action="edit", path="big.xlsx", append_rows=rows)
        assert "error" in result


@_needs_openpyxl
class TestPathValidation:
    @pytest.mark.asyncio
    async def test_blocked_system_path_rejected(self):
        result = await handle(
            action="create",
            path="/etc/shadow",
            sheets=[{"name": "Sheet1", "rows": [["x"]]}],
        )
        assert "error" in result

    @pytest.mark.asyncio
    async def test_null_bytes_rejected(self):
        result = await handle(action="read", path="test\x00.xlsx")
        assert "error" in result


@_needs_openpyxl
class TestUnknownAction:
    @pytest.mark.asyncio
    async def test_unknown_action(self):
        result = await handle(action="delete", path="test.xlsx")
        assert "error" in result
        assert "Unknown action" in result["error"]


class TestGracefulDegradation:
    @pytest.mark.asyncio
    async def test_unavailable(self):
        with patch("anteroom.tools.office_xlsx.AVAILABLE", False):
            result = await handle(action="read", path="test.xlsx")
            assert "error" in result
            assert "pip install" in result["error"]


# ---------------------------------------------------------------------------
# Tests for new actions (lib backend)
# ---------------------------------------------------------------------------


async def _create_test_workbook(filename: str, rows: list | None = None, headers: list | None = None) -> dict:
    """Helper to create a test workbook for use in subsequent action tests."""
    sheets = [{"name": "Data", "rows": rows or [["a", "b"], ["c", "d"]], "headers": headers or []}]
    return await handle(action="create", path=filename, sheets=sheets)


@_needs_openpyxl
class TestFormatCells:
    @pytest.mark.asyncio
    async def test_format_font(self, tmp_path):
        await _create_test_workbook("fmt.xlsx")
        result = await handle(
            action="format_cells",
            path="fmt.xlsx",
            cell_range="A1:B1",
            format={"font": {"name": "Arial", "size": 14, "bold": True, "italic": True, "color": "#FF0000"}},
        )
        assert "error" not in result
        assert "Formatted" in result["result"]
        assert result["path"] == "fmt.xlsx"

    @pytest.mark.asyncio
    async def test_format_fill(self, tmp_path):
        await _create_test_workbook("fmt_fill.xlsx")
        result = await handle(
            action="format_cells",
            path="fmt_fill.xlsx",
            cell_range="A1:A1",
            format={"fill": {"color": "#00FF00"}},
        )
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_format_alignment(self, tmp_path):
        await _create_test_workbook("fmt_align.xlsx")
        result = await handle(
            action="format_cells",
            path="fmt_align.xlsx",
            cell_range="A1:B2",
            format={"alignment": {"horizontal": "center", "vertical": "top", "wrap_text": True}},
        )
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_format_number_format(self, tmp_path):
        await _create_test_workbook("fmt_num.xlsx", rows=[[1.5, 2.7]])
        result = await handle(
            action="format_cells",
            path="fmt_num.xlsx",
            cell_range="A1:B1",
            format={"number_format": "0.00%"},
        )
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_format_missing_range(self, tmp_path):
        await _create_test_workbook("fmt_no_range.xlsx")
        result = await handle(action="format_cells", path="fmt_no_range.xlsx")
        assert "error" in result
        assert "cell_range" in result["error"]

    @pytest.mark.asyncio
    async def test_format_file_not_found(self):
        result = await handle(action="format_cells", path="missing.xlsx", cell_range="A1")
        assert "error" in result

    @pytest.mark.asyncio
    async def test_format_combined(self, tmp_path):
        await _create_test_workbook("fmt_combo.xlsx")
        result = await handle(
            action="format_cells",
            path="fmt_combo.xlsx",
            cell_range="A1:A1",
            format={
                "font": {"bold": True, "color": "#0000FF"},
                "fill": {"color": "#FFFF00"},
                "alignment": {"horizontal": "right"},
                "number_format": "#,##0",
            },
        )
        assert "error" not in result


@_needs_openpyxl
class TestMergeCells:
    @pytest.mark.asyncio
    async def test_merge(self, tmp_path):
        await _create_test_workbook("merge.xlsx")
        result = await handle(action="merge_cells", path="merge.xlsx", cell_range="A1:B1", merge=True)
        assert "error" not in result
        assert "Merged" in result["result"]

    @pytest.mark.asyncio
    async def test_unmerge(self, tmp_path):
        await _create_test_workbook("unmerge.xlsx")
        await handle(action="merge_cells", path="unmerge.xlsx", cell_range="A1:B1", merge=True)
        result = await handle(action="merge_cells", path="unmerge.xlsx", cell_range="A1:B1", merge=False)
        assert "error" not in result
        assert "Unmerged" in result["result"]

    @pytest.mark.asyncio
    async def test_merge_default_is_true(self, tmp_path):
        await _create_test_workbook("merge_def.xlsx")
        result = await handle(action="merge_cells", path="merge_def.xlsx", cell_range="A1:B2")
        assert "error" not in result
        assert "Merged" in result["result"]

    @pytest.mark.asyncio
    async def test_merge_missing_range(self, tmp_path):
        await _create_test_workbook("merge_no_range.xlsx")
        result = await handle(action="merge_cells", path="merge_no_range.xlsx")
        assert "error" in result
        assert "cell_range" in result["error"]

    @pytest.mark.asyncio
    async def test_merge_file_not_found(self):
        result = await handle(action="merge_cells", path="missing.xlsx", cell_range="A1:B1")
        assert "error" in result


@_needs_openpyxl
class TestFreezePanes:
    @pytest.mark.asyncio
    async def test_freeze_default(self, tmp_path):
        await _create_test_workbook("freeze.xlsx")
        result = await handle(action="freeze_panes", path="freeze.xlsx")
        assert "error" not in result
        assert "row 2" in result["result"]
        assert "column 1" in result["result"]

    @pytest.mark.asyncio
    async def test_freeze_custom(self, tmp_path):
        await _create_test_workbook("freeze_custom.xlsx")
        result = await handle(action="freeze_panes", path="freeze_custom.xlsx", row=3, column=2)
        assert "error" not in result
        assert "row 3" in result["result"]
        assert "column 2" in result["result"]

    @pytest.mark.asyncio
    async def test_freeze_file_not_found(self):
        result = await handle(action="freeze_panes", path="missing.xlsx")
        assert "error" in result

    @pytest.mark.asyncio
    async def test_freeze_persists(self, tmp_path):
        await _create_test_workbook("freeze_check.xlsx")
        await handle(action="freeze_panes", path="freeze_check.xlsx", row=2, column=1)

        import openpyxl

        wb = openpyxl.load_workbook(str(tmp_path / "freeze_check.xlsx"))
        ws = wb.active
        assert ws.freeze_panes == "A2"
        wb.close()


@_needs_openpyxl
class TestAutoFilter:
    @pytest.mark.asyncio
    async def test_enable_default(self, tmp_path):
        await _create_test_workbook("af.xlsx")
        result = await handle(action="auto_filter", path="af.xlsx")
        assert "error" not in result
        assert "Auto-filter updated" in result["result"]

    @pytest.mark.asyncio
    async def test_enable_with_range(self, tmp_path):
        await _create_test_workbook("af_range.xlsx")
        result = await handle(action="auto_filter", path="af_range.xlsx", cell_range="A1:B2", operation="enable")
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_disable(self, tmp_path):
        await _create_test_workbook("af_dis.xlsx")
        await handle(action="auto_filter", path="af_dis.xlsx", operation="enable")
        result = await handle(action="auto_filter", path="af_dis.xlsx", operation="disable")
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_auto_filter_file_not_found(self):
        result = await handle(action="auto_filter", path="missing.xlsx")
        assert "error" in result


@_needs_openpyxl
class TestPrintArea:
    @pytest.mark.asyncio
    async def test_set_print_area(self, tmp_path):
        await _create_test_workbook("pa.xlsx")
        result = await handle(action="print_area", path="pa.xlsx", cell_range="A1:B5")
        assert "error" not in result
        assert "print area" in result["result"].lower()

    @pytest.mark.asyncio
    async def test_print_area_persists(self, tmp_path):
        await _create_test_workbook("pa_check.xlsx")
        await handle(action="print_area", path="pa_check.xlsx", cell_range="A1:C10")

        import openpyxl

        wb = openpyxl.load_workbook(str(tmp_path / "pa_check.xlsx"))
        ws = wb.active
        assert ws.print_area is not None
        pa_str = str(ws.print_area)
        assert "A" in pa_str and "C" in pa_str and "1" in pa_str and "10" in pa_str
        wb.close()

    @pytest.mark.asyncio
    async def test_print_area_missing_range(self, tmp_path):
        await _create_test_workbook("pa_no_range.xlsx")
        result = await handle(action="print_area", path="pa_no_range.xlsx")
        assert "error" in result
        assert "cell_range" in result["error"]

    @pytest.mark.asyncio
    async def test_print_area_file_not_found(self):
        result = await handle(action="print_area", path="missing.xlsx", cell_range="A1:B5")
        assert "error" in result


@_needs_openpyxl
class TestNamedRanges:
    @pytest.mark.asyncio
    async def test_add_named_range(self, tmp_path):
        await _create_test_workbook("nr.xlsx")
        result = await handle(
            action="named_ranges",
            path="nr.xlsx",
            operation="add",
            range_name="MyRange",
            cell_range="A1:B2",
        )
        assert "error" not in result
        assert "Added" in result["result"]

    @pytest.mark.asyncio
    async def test_list_named_ranges(self, tmp_path):
        await _create_test_workbook("nr_list.xlsx")
        await handle(
            action="named_ranges",
            path="nr_list.xlsx",
            operation="add",
            range_name="TestRange",
            cell_range="A1:B2",
        )
        result = await handle(action="named_ranges", path="nr_list.xlsx", operation="list")
        assert "error" not in result
        assert len(result["names"]) >= 1
        names = [n["name"] for n in result["names"]]
        assert "TestRange" in names

    @pytest.mark.asyncio
    async def test_delete_named_range(self, tmp_path):
        await _create_test_workbook("nr_del.xlsx")
        await handle(
            action="named_ranges",
            path="nr_del.xlsx",
            operation="add",
            range_name="DeleteMe",
            cell_range="A1:B1",
        )
        result = await handle(
            action="named_ranges",
            path="nr_del.xlsx",
            operation="delete",
            range_name="DeleteMe",
        )
        assert "error" not in result
        assert "Deleted" in result["result"]

    @pytest.mark.asyncio
    async def test_add_missing_params(self, tmp_path):
        await _create_test_workbook("nr_err.xlsx")
        result = await handle(action="named_ranges", path="nr_err.xlsx", operation="add")
        assert "error" in result

    @pytest.mark.asyncio
    async def test_delete_missing_name(self, tmp_path):
        await _create_test_workbook("nr_err2.xlsx")
        result = await handle(action="named_ranges", path="nr_err2.xlsx", operation="delete")
        assert "error" in result

    @pytest.mark.asyncio
    async def test_unknown_operation(self, tmp_path):
        await _create_test_workbook("nr_unk.xlsx")
        result = await handle(action="named_ranges", path="nr_unk.xlsx", operation="bogus")
        assert "error" in result

    @pytest.mark.asyncio
    async def test_list_empty(self, tmp_path):
        await _create_test_workbook("nr_empty.xlsx")
        result = await handle(action="named_ranges", path="nr_empty.xlsx", operation="list")
        assert "error" not in result
        assert result["names"] == []


@_needs_openpyxl
class TestDataValidation:
    @pytest.mark.asyncio
    async def test_add_list_validation(self, tmp_path):
        await _create_test_workbook("dv.xlsx")
        result = await handle(
            action="data_validation",
            path="dv.xlsx",
            cell_range="A1:A10",
            validation={"type": "list", "formula1": '"Yes,No,Maybe"'},
        )
        assert "error" not in result
        assert "data validation" in result["result"].lower()

    @pytest.mark.asyncio
    async def test_add_whole_validation(self, tmp_path):
        await _create_test_workbook("dv_whole.xlsx")
        result = await handle(
            action="data_validation",
            path="dv_whole.xlsx",
            cell_range="B1:B5",
            validation={"type": "whole", "formula1": "1", "operator": "greaterThan"},
        )
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_validation_with_error_message(self, tmp_path):
        await _create_test_workbook("dv_err.xlsx")
        result = await handle(
            action="data_validation",
            path="dv_err.xlsx",
            cell_range="A1",
            validation={"type": "list", "formula1": '"A,B,C"', "error_message": "Pick one"},
        )
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_validation_missing_params(self, tmp_path):
        await _create_test_workbook("dv_miss.xlsx")
        result = await handle(action="data_validation", path="dv_miss.xlsx")
        assert "error" in result

    @pytest.mark.asyncio
    async def test_validation_missing_range(self, tmp_path):
        await _create_test_workbook("dv_miss2.xlsx")
        result = await handle(
            action="data_validation",
            path="dv_miss2.xlsx",
            validation={"type": "list", "formula1": '"X"'},
        )
        assert "error" in result


@_needs_openpyxl
class TestConditionalFormat:
    @pytest.mark.asyncio
    async def test_add_conditional_format(self, tmp_path):
        await _create_test_workbook("cf.xlsx", rows=[[10], [20], [30]])
        result = await handle(
            action="conditional_format",
            path="cf.xlsx",
            cell_range="A1:A3",
            rule={"operator": "greaterThan", "formula": "15"},
        )
        assert "error" not in result
        assert "conditional format" in result["result"].lower()

    @pytest.mark.asyncio
    async def test_conditional_format_with_styling(self, tmp_path):
        await _create_test_workbook("cf_style.xlsx", rows=[[5], [25]])
        result = await handle(
            action="conditional_format",
            path="cf_style.xlsx",
            cell_range="A1:A2",
            rule={
                "operator": "greaterThan",
                "formula": "10",
                "format": {"font_color": "#FF0000", "fill_color": "#FFFF00"},
            },
        )
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_conditional_format_missing_params(self, tmp_path):
        await _create_test_workbook("cf_miss.xlsx")
        result = await handle(action="conditional_format", path="cf_miss.xlsx")
        assert "error" in result

    @pytest.mark.asyncio
    async def test_conditional_format_missing_rule(self, tmp_path):
        await _create_test_workbook("cf_miss2.xlsx")
        result = await handle(action="conditional_format", path="cf_miss2.xlsx", cell_range="A1")
        assert "error" in result


@_needs_openpyxl
class TestComments:
    @pytest.mark.asyncio
    async def test_add_comment(self, tmp_path):
        await _create_test_workbook("cmt.xlsx")
        result = await handle(
            action="comments",
            path="cmt.xlsx",
            operation="add",
            cell_range="A1",
            comment_text="This is a test comment",
        )
        assert "error" not in result
        assert "Added comment" in result["result"]

    @pytest.mark.asyncio
    async def test_add_comment_with_author(self, tmp_path):
        await _create_test_workbook("cmt_auth.xlsx")
        result = await handle(
            action="comments",
            path="cmt_auth.xlsx",
            operation="add",
            cell_range="A1",
            comment_text="Author test",
            author="TestUser",
        )
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_read_comments(self, tmp_path):
        await _create_test_workbook("cmt_read.xlsx")
        await handle(
            action="comments",
            path="cmt_read.xlsx",
            operation="add",
            cell_range="A1",
            comment_text="Hello",
        )
        result = await handle(action="comments", path="cmt_read.xlsx", operation="read")
        assert "error" not in result
        assert len(result["comments"]) >= 1
        assert result["comments"][0]["text"] == "Hello"

    @pytest.mark.asyncio
    async def test_delete_comment(self, tmp_path):
        await _create_test_workbook("cmt_del.xlsx")
        await handle(
            action="comments",
            path="cmt_del.xlsx",
            operation="add",
            cell_range="A1",
            comment_text="Delete me",
        )
        result = await handle(
            action="comments",
            path="cmt_del.xlsx",
            operation="delete",
            cell_range="A1",
        )
        assert "error" not in result
        assert "Deleted" in result["result"]

        read_result = await handle(action="comments", path="cmt_del.xlsx", operation="read")
        assert len(read_result["comments"]) == 0

    @pytest.mark.asyncio
    async def test_add_comment_missing_params(self, tmp_path):
        await _create_test_workbook("cmt_miss.xlsx")
        result = await handle(action="comments", path="cmt_miss.xlsx", operation="add")
        assert "error" in result

    @pytest.mark.asyncio
    async def test_delete_comment_missing_range(self, tmp_path):
        await _create_test_workbook("cmt_miss2.xlsx")
        result = await handle(action="comments", path="cmt_miss2.xlsx", operation="delete")
        assert "error" in result

    @pytest.mark.asyncio
    async def test_unknown_operation(self, tmp_path):
        await _create_test_workbook("cmt_unk.xlsx")
        result = await handle(action="comments", path="cmt_unk.xlsx", operation="bogus")
        assert "error" in result

    @pytest.mark.asyncio
    async def test_read_empty_comments(self, tmp_path):
        await _create_test_workbook("cmt_empty.xlsx")
        result = await handle(action="comments", path="cmt_empty.xlsx", operation="read")
        assert "error" not in result
        assert result["comments"] == []


@_needs_openpyxl
class TestHyperlinks:
    @pytest.mark.asyncio
    async def test_add_hyperlink(self, tmp_path):
        await _create_test_workbook("hl.xlsx")
        result = await handle(
            action="hyperlinks",
            path="hl.xlsx",
            operation="add",
            cell_range="A1",
            url="https://example.com",
        )
        assert "error" not in result
        assert "Added hyperlink" in result["result"]

    @pytest.mark.asyncio
    async def test_add_hyperlink_with_display_text(self, tmp_path):
        await _create_test_workbook("hl_disp.xlsx")
        result = await handle(
            action="hyperlinks",
            path="hl_disp.xlsx",
            operation="add",
            cell_range="A1",
            url="https://example.com",
            display_text="Example Site",
        )
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_read_hyperlinks(self, tmp_path):
        await _create_test_workbook("hl_read.xlsx")
        await handle(
            action="hyperlinks",
            path="hl_read.xlsx",
            operation="add",
            cell_range="A1",
            url="https://example.com",
            display_text="Example",
        )
        result = await handle(action="hyperlinks", path="hl_read.xlsx", operation="read")
        assert "error" not in result
        assert len(result["hyperlinks"]) >= 1
        assert "example.com" in result["hyperlinks"][0]["url"]

    @pytest.mark.asyncio
    async def test_add_missing_params(self, tmp_path):
        await _create_test_workbook("hl_miss.xlsx")
        result = await handle(action="hyperlinks", path="hl_miss.xlsx", operation="add")
        assert "error" in result

    @pytest.mark.asyncio
    async def test_unknown_operation(self, tmp_path):
        await _create_test_workbook("hl_unk.xlsx")
        result = await handle(action="hyperlinks", path="hl_unk.xlsx", operation="bogus")
        assert "error" in result

    @pytest.mark.asyncio
    async def test_read_empty_hyperlinks(self, tmp_path):
        await _create_test_workbook("hl_empty.xlsx")
        result = await handle(action="hyperlinks", path="hl_empty.xlsx", operation="read")
        assert "error" not in result
        assert result["hyperlinks"] == []


@_needs_openpyxl
class TestImages:
    @pytest.fixture
    def small_png(self, tmp_path):
        """Create a minimal valid 1x1 PNG file."""
        import struct
        import zlib

        def _chunk(chunk_type: bytes, data: bytes) -> bytes:
            raw = chunk_type + data
            return struct.pack(">I", len(data)) + raw + struct.pack(">I", zlib.crc32(raw) & 0xFFFFFFFF)

        sig = b"\x89PNG\r\n\x1a\n"
        ihdr = struct.pack(">IIBBBBB", 1, 1, 8, 2, 0, 0, 0)
        scanline = b"\x00\xff\x00\x00"
        idat = zlib.compress(scanline)
        png_bytes = sig + _chunk(b"IHDR", ihdr) + _chunk(b"IDAT", idat) + _chunk(b"IEND", b"")
        png_path = tmp_path / "test_image.png"
        png_path.write_bytes(png_bytes)
        return png_path

    @pytest.mark.asyncio
    async def test_insert_image(self, tmp_path, small_png):
        await _create_test_workbook("img.xlsx")
        result = await handle(
            action="images",
            path="img.xlsx",
            image_path="test_image.png",
            anchor_cell="C3",
        )
        assert "error" not in result
        assert "Inserted image" in result["result"]
        assert "C3" in result["result"]

    @pytest.mark.asyncio
    async def test_insert_image_default_anchor(self, tmp_path, small_png):
        await _create_test_workbook("img_def.xlsx")
        result = await handle(action="images", path="img_def.xlsx", image_path="test_image.png")
        assert "error" not in result
        assert "A1" in result["result"]

    @pytest.mark.asyncio
    async def test_image_missing_path(self, tmp_path):
        await _create_test_workbook("img_miss.xlsx")
        result = await handle(action="images", path="img_miss.xlsx")
        assert "error" in result
        assert "image_path" in result["error"]

    @pytest.mark.asyncio
    async def test_image_not_found(self, tmp_path):
        await _create_test_workbook("img_nf.xlsx")
        result = await handle(action="images", path="img_nf.xlsx", image_path="nonexistent.png")
        assert "error" in result
        assert "not found" in result["error"].lower()

    @pytest.mark.asyncio
    async def test_image_file_not_found(self):
        result = await handle(action="images", path="missing.xlsx", image_path="test.png")
        assert "error" in result


@_needs_openpyxl
class TestProtect:
    @pytest.mark.asyncio
    async def test_enable_sheet_protection(self, tmp_path):
        await _create_test_workbook("prot.xlsx")
        result = await handle(
            action="protect",
            path="prot.xlsx",
            operation="enable",
            sheet_name="Data",
        )
        assert "error" not in result
        assert "Protected" in result["result"]

    @pytest.mark.asyncio
    async def test_enable_workbook_protection(self, tmp_path):
        await _create_test_workbook("prot_wb.xlsx")
        result = await handle(action="protect", path="prot_wb.xlsx", operation="enable")
        assert "error" not in result
        assert "Protected" in result["result"]

    @pytest.mark.asyncio
    async def test_enable_with_password(self, tmp_path):
        await _create_test_workbook("prot_pw.xlsx")
        result = await handle(
            action="protect",
            path="prot_pw.xlsx",
            operation="enable",
            sheet_name="Data",
            password="secret123",
        )
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_disable_sheet_protection(self, tmp_path):
        await _create_test_workbook("prot_dis.xlsx")
        await handle(action="protect", path="prot_dis.xlsx", operation="enable", sheet_name="Data")
        result = await handle(action="protect", path="prot_dis.xlsx", operation="disable", sheet_name="Data")
        assert "error" not in result
        assert "Removed protection" in result["result"]

    @pytest.mark.asyncio
    async def test_disable_workbook_protection(self, tmp_path):
        await _create_test_workbook("prot_dis_wb.xlsx")
        await handle(action="protect", path="prot_dis_wb.xlsx", operation="enable")
        result = await handle(action="protect", path="prot_dis_wb.xlsx", operation="disable")
        assert "error" not in result
        assert "Removed protection" in result["result"]

    @pytest.mark.asyncio
    async def test_protect_file_not_found(self):
        result = await handle(action="protect", path="missing.xlsx", operation="enable")
        assert "error" in result

    @pytest.mark.asyncio
    async def test_protection_persists(self, tmp_path):
        await _create_test_workbook("prot_check.xlsx")
        await handle(action="protect", path="prot_check.xlsx", operation="enable", sheet_name="Data")

        import openpyxl

        wb = openpyxl.load_workbook(str(tmp_path / "prot_check.xlsx"))
        ws = wb["Data"]
        assert ws.protection.sheet is True
        wb.close()


@_needs_openpyxl
class TestGroupRowsCols:
    @pytest.mark.asyncio
    async def test_group_rows(self, tmp_path):
        await _create_test_workbook("grp.xlsx", rows=[[i] for i in range(10)])
        result = await handle(
            action="group_rows_cols",
            path="grp.xlsx",
            start=2,
            end=5,
            axis="rows",
            operation="group",
        )
        assert "error" not in result
        assert "Grouped" in result["result"]
        assert "rows" in result["result"]

    @pytest.mark.asyncio
    async def test_ungroup_rows(self, tmp_path):
        await _create_test_workbook("grp_un.xlsx", rows=[[i] for i in range(10)])
        await handle(action="group_rows_cols", path="grp_un.xlsx", start=2, end=5, axis="rows", operation="group")
        result = await handle(
            action="group_rows_cols",
            path="grp_un.xlsx",
            start=2,
            end=5,
            axis="rows",
            operation="ungroup",
        )
        assert "error" not in result
        assert "Ungrouped" in result["result"]

    @pytest.mark.asyncio
    async def test_group_columns(self, tmp_path):
        await _create_test_workbook("grp_col.xlsx")
        result = await handle(
            action="group_rows_cols",
            path="grp_col.xlsx",
            start=1,
            end=3,
            axis="columns",
            operation="group",
        )
        assert "error" not in result
        assert "columns" in result["result"]

    @pytest.mark.asyncio
    async def test_group_missing_params(self, tmp_path):
        await _create_test_workbook("grp_miss.xlsx")
        result = await handle(action="group_rows_cols", path="grp_miss.xlsx", axis="rows")
        assert "error" in result
        assert "start" in result["error"]

    @pytest.mark.asyncio
    async def test_group_default_axis(self, tmp_path):
        await _create_test_workbook("grp_def.xlsx", rows=[[i] for i in range(10)])
        result = await handle(action="group_rows_cols", path="grp_def.xlsx", start=1, end=3)
        assert "error" not in result
        assert "rows" in result["result"]


@_needs_openpyxl
class TestPrintSettings:
    @pytest.mark.asyncio
    async def test_set_orientation(self, tmp_path):
        await _create_test_workbook("ps.xlsx")
        result = await handle(
            action="print_settings",
            path="ps.xlsx",
            page_setup={"orientation": "landscape"},
        )
        assert "error" not in result
        assert "print settings" in result["result"].lower()

    @pytest.mark.asyncio
    async def test_set_margins(self, tmp_path):
        await _create_test_workbook("ps_margins.xlsx")
        result = await handle(
            action="print_settings",
            path="ps_margins.xlsx",
            page_setup={"margins": {"top": 1.0, "bottom": 1.0, "left": 0.5, "right": 0.5}},
        )
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_set_fit_to_page(self, tmp_path):
        await _create_test_workbook("ps_fit.xlsx")
        result = await handle(
            action="print_settings",
            path="ps_fit.xlsx",
            page_setup={"fit_to_width": 1, "fit_to_height": 1},
        )
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_set_header_footer(self, tmp_path):
        await _create_test_workbook("ps_hf.xlsx")
        result = await handle(
            action="print_settings",
            path="ps_hf.xlsx",
            page_setup={"header": "My Report", "footer": "Page &P"},
        )
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_set_paper_size(self, tmp_path):
        await _create_test_workbook("ps_paper.xlsx")
        result = await handle(
            action="print_settings",
            path="ps_paper.xlsx",
            page_setup={"paper_size": 9},
        )
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_combined_settings(self, tmp_path):
        await _create_test_workbook("ps_combo.xlsx")
        result = await handle(
            action="print_settings",
            path="ps_combo.xlsx",
            page_setup={
                "orientation": "portrait",
                "fit_to_width": 1,
                "margins": {"top": 0.75, "bottom": 0.75},
                "header": "Title",
            },
        )
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_print_settings_file_not_found(self):
        result = await handle(action="print_settings", path="missing.xlsx", page_setup={"orientation": "landscape"})
        assert "error" in result

    @pytest.mark.asyncio
    async def test_orientation_persists(self, tmp_path):
        await _create_test_workbook("ps_check.xlsx")
        await handle(action="print_settings", path="ps_check.xlsx", page_setup={"orientation": "landscape"})

        import openpyxl

        wb = openpyxl.load_workbook(str(tmp_path / "ps_check.xlsx"))
        ws = wb.active
        assert ws.page_setup.orientation == "landscape"
        wb.close()


@_needs_openpyxl
class TestCharts:
    @pytest.mark.asyncio
    async def test_create_column_chart(self, tmp_path):
        await handle(
            action="create",
            path="chart.xlsx",
            sheets=[
                {
                    "name": "Data",
                    "headers": ["Month", "Sales"],
                    "rows": [["Jan", 100], ["Feb", 200], ["Mar", 150]],
                }
            ],
        )
        result = await handle(
            action="charts",
            path="chart.xlsx",
            data_range="A1:B4",
            chart_type="column",
            chart_title="Monthly Sales",
        )
        assert "error" not in result
        assert "column chart" in result["result"].lower()

    @pytest.mark.asyncio
    async def test_create_line_chart(self, tmp_path):
        await handle(
            action="create",
            path="chart_line.xlsx",
            sheets=[{"name": "Data", "headers": ["X", "Y"], "rows": [[1, 10], [2, 20], [3, 30]]}],
        )
        result = await handle(action="charts", path="chart_line.xlsx", data_range="A1:B4", chart_type="line")
        assert "error" not in result
        assert "line" in result["result"]

    @pytest.mark.asyncio
    async def test_create_pie_chart(self, tmp_path):
        await handle(
            action="create",
            path="chart_pie.xlsx",
            sheets=[{"name": "Data", "headers": ["Category", "Value"], "rows": [["A", 30], ["B", 50], ["C", 20]]}],
        )
        result = await handle(action="charts", path="chart_pie.xlsx", data_range="A1:B4", chart_type="pie")
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_create_bar_chart(self, tmp_path):
        await handle(
            action="create",
            path="chart_bar.xlsx",
            sheets=[{"name": "Data", "rows": [["A", 10], ["B", 20]]}],
        )
        result = await handle(action="charts", path="chart_bar.xlsx", data_range="A1:B2", chart_type="bar")
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_create_scatter_chart(self, tmp_path):
        await handle(
            action="create",
            path="chart_scatter.xlsx",
            sheets=[{"name": "Data", "rows": [[1, 2], [3, 4], [5, 6]]}],
        )
        result = await handle(action="charts", path="chart_scatter.xlsx", data_range="A1:B3", chart_type="scatter")
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_create_area_chart(self, tmp_path):
        await handle(
            action="create",
            path="chart_area.xlsx",
            sheets=[{"name": "Data", "rows": [[1, 5], [2, 10], [3, 8]]}],
        )
        result = await handle(action="charts", path="chart_area.xlsx", data_range="A1:B3", chart_type="area")
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_chart_custom_anchor(self, tmp_path):
        await handle(
            action="create",
            path="chart_anchor.xlsx",
            sheets=[{"name": "Data", "rows": [[1, 2], [3, 4]]}],
        )
        result = await handle(
            action="charts",
            path="chart_anchor.xlsx",
            data_range="A1:B2",
            anchor_cell="F5",
        )
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_chart_missing_data_range(self, tmp_path):
        await _create_test_workbook("chart_miss.xlsx")
        result = await handle(action="charts", path="chart_miss.xlsx")
        assert "error" in result
        assert "data_range" in result["error"]

    @pytest.mark.asyncio
    async def test_chart_file_not_found(self):
        result = await handle(action="charts", path="missing.xlsx", data_range="A1:B3")
        assert "error" in result


# ---------------------------------------------------------------------------
# Security: DDE formula injection sanitization
# ---------------------------------------------------------------------------


class TestSanitizeCellValue:
    def test_normal_string_passes_through(self):
        assert _sanitize_cell_value("hello") == "hello"

    def test_number_passes_through(self):
        assert _sanitize_cell_value(42) == 42

    def test_none_passes_through(self):
        assert _sanitize_cell_value(None) is None

    def test_legitimate_formula_passes_through(self):
        assert _sanitize_cell_value("=SUM(A1:A10)") == "=SUM(A1:A10)"

    def test_dde_cmd_blocked(self):
        val = "=CMD|'/C calc'!A0"
        result = _sanitize_cell_value(val)
        assert result.startswith("'")

    def test_dde_msexcel_blocked(self):
        val = "=MSEXCEL|'\\..\\file'!A0"
        result = _sanitize_cell_value(val)
        assert result.startswith("'")

    def test_dde_with_plus_prefix(self):
        val = "+CMD|'/C whoami'!A0"
        result = _sanitize_cell_value(val)
        assert result.startswith("'")

    def test_dde_with_minus_prefix(self):
        val = "-CMD|'/C id'!A0"
        result = _sanitize_cell_value(val)
        assert result.startswith("'")

    def test_dde_with_at_prefix(self):
        val = "@SUM|foo!A0"
        result = _sanitize_cell_value(val)
        assert result.startswith("'")


# ---------------------------------------------------------------------------
# COM-only actions (should return COM-only error on lib backend)
# ---------------------------------------------------------------------------


@_needs_openpyxl
class TestExportPdfComOnly:
    @pytest.mark.asyncio
    async def test_returns_com_only_error(self, tmp_path):
        await _create_test_workbook("pdf.xlsx")
        result = await handle(action="export_pdf", path="pdf.xlsx")
        assert "error" in result
        assert "COM backend" in result["error"]


@_needs_openpyxl
class TestSortComOnly:
    @pytest.mark.asyncio
    async def test_returns_com_only_error(self, tmp_path):
        await _create_test_workbook("sort.xlsx")
        result = await handle(action="sort", path="sort.xlsx", cell_range="A1:B3", sort_column="A")
        assert "error" in result
        assert "COM backend" in result["error"]


@_needs_openpyxl
class TestPivotTablesComOnly:
    @pytest.mark.asyncio
    async def test_returns_com_only_error(self, tmp_path):
        await _create_test_workbook("pivot.xlsx")
        result = await handle(action="pivot_tables", path="pivot.xlsx", data_range="A1:B3")
        assert "error" in result
        assert "COM backend" in result["error"]


@_needs_openpyxl
class TestSparklinesComOnly:
    @pytest.mark.asyncio
    async def test_returns_com_only_error(self, tmp_path):
        await _create_test_workbook("spark.xlsx")
        result = await handle(action="sparklines", path="spark.xlsx", data_range="A1:D1", anchor_cell="E1")
        assert "error" in result
        assert "COM backend" in result["error"]


@_needs_openpyxl
class TestSlicersComOnly:
    @pytest.mark.asyncio
    async def test_returns_com_only_error(self, tmp_path):
        await _create_test_workbook("slicer.xlsx")
        result = await handle(action="slicers", path="slicer.xlsx", range_name="PivotTable1")
        assert "error" in result
        assert "COM backend" in result["error"]


class TestComDispatchErrorHandling:
    @pytest.mark.asyncio
    async def test_dispatch_com_returns_error_dict_on_exception(self):
        from unittest.mock import AsyncMock, MagicMock

        mock_manager = MagicMock()
        mock_manager.run_com = AsyncMock(side_effect=RuntimeError("Access denied by security policy"))
        mock_com_mod = MagicMock()
        mock_com_mod.get_manager.return_value = mock_manager
        mock_com_mod.COM_AVAILABLE = True

        with (
            patch("anteroom.tools.office_xlsx.AVAILABLE", True),
            patch("anteroom.tools.office_xlsx._BACKEND", "com"),
            patch("anteroom.tools.office_xlsx._com_mod", mock_com_mod),
        ):
            result = await handle(action="edit", path="test.xlsx", cells=[{"cell": "A1", "value": "x"}])

        assert "error" in result
        assert "Access denied by security policy" in result["error"]
        assert "RuntimeError" in result["error"]


# ---------------------------------------------------------------------------
# Enriched read output tests
# ---------------------------------------------------------------------------


@_needs_openpyxl
class TestEnrichedRead:
    @pytest.mark.asyncio
    async def test_read_returns_formulas(self, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = 10
        ws["A2"] = 20
        ws["A3"] = "=SUM(A1:A2)"
        wb.save(str(tmp_path / "formulas.xlsx"))
        wb.close()
        result = await handle(action="read", path="formulas.xlsx")
        assert "error" not in result
        assert "formulas" in result
        assert "A3" in result["formulas"]
        assert result["formulas"]["A3"] == "=SUM(A1:A2)"

    @pytest.mark.asyncio
    async def test_read_returns_merged_ranges(self, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "merged"
        ws.merge_cells("A1:C1")
        wb.save(str(tmp_path / "merged.xlsx"))
        wb.close()
        result = await handle(action="read", path="merged.xlsx")
        assert "error" not in result
        assert "merged_ranges" in result
        assert "A1:C1" in result["merged_ranges"]

    @pytest.mark.asyncio
    async def test_read_returns_formatting(self, tmp_path):
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "bold text"
        ws["A1"].font = Font(bold=True, size=14)
        wb.save(str(tmp_path / "fmt.xlsx"))
        wb.close()
        result = await handle(action="read", path="fmt.xlsx")
        assert "error" not in result
        assert "formatting" in result
        assert "A1" in result["formatting"]
        assert result["formatting"]["A1"]["bold"] is True
        assert result["formatting"]["A1"]["size"] == 14

    @pytest.mark.asyncio
    async def test_read_returns_column_widths(self, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "wide"
        ws.column_dimensions["A"].width = 25.0
        wb.save(str(tmp_path / "widths.xlsx"))
        wb.close()
        result = await handle(action="read", path="widths.xlsx")
        assert "error" not in result
        assert "column_widths" in result
        assert result["column_widths"]["A"] == 25.0

    @pytest.mark.asyncio
    async def test_read_returns_row_heights(self, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "tall"
        ws.row_dimensions[1].height = 30.0
        wb.save(str(tmp_path / "heights.xlsx"))
        wb.close()
        result = await handle(action="read", path="heights.xlsx")
        assert "error" not in result
        assert "row_heights" in result
        assert result["row_heights"][1] == 30.0

    @pytest.mark.asyncio
    async def test_read_returns_freeze_pane(self, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "data"
        ws.freeze_panes = "B2"
        wb.save(str(tmp_path / "freeze.xlsx"))
        wb.close()
        result = await handle(action="read", path="freeze.xlsx")
        assert "error" not in result
        assert "freeze_pane" in result
        assert result["freeze_pane"] == "B2"

    @pytest.mark.asyncio
    async def test_read_returns_auto_filter(self, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Age"])
        ws.append(["Alice", 30])
        ws.auto_filter.ref = "A1:B2"
        wb.save(str(tmp_path / "filter.xlsx"))
        wb.close()
        result = await handle(action="read", path="filter.xlsx")
        assert "error" not in result
        assert "auto_filter" in result
        assert result["auto_filter"] == "A1:B2"

    @pytest.mark.asyncio
    async def test_read_returns_named_ranges(self, tmp_path):
        import openpyxl
        from openpyxl.workbook.defined_name import DefinedName

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws["A1"] = "value"
        dn = DefinedName("my_range", attr_text="'Data'!$A$1")
        wb.defined_names.add(dn)
        wb.save(str(tmp_path / "named.xlsx"))
        wb.close()
        result = await handle(action="read", path="named.xlsx")
        assert "error" not in result
        assert "named_ranges" in result
        assert any(nr["name"] == "my_range" for nr in result["named_ranges"])

    @pytest.mark.asyncio
    async def test_read_returns_data_validations(self, tmp_path):
        import openpyxl
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "pick"
        dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
        dv.sqref = "A1"
        ws.add_data_validation(dv)
        wb.save(str(tmp_path / "validation.xlsx"))
        wb.close()
        result = await handle(action="read", path="validation.xlsx")
        assert "error" not in result
        assert "data_validations" in result
        assert len(result["data_validations"]) >= 1
        assert result["data_validations"][0]["type"] == "list"

    @pytest.mark.asyncio
    async def test_read_returns_hidden_sheets(self, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Visible"
        ws1["A1"] = "visible"
        ws2 = wb.create_sheet("Hidden")
        ws2["A1"] = "secret"
        ws2.sheet_state = "hidden"
        wb.save(str(tmp_path / "hidden.xlsx"))
        wb.close()
        result = await handle(action="read", path="hidden.xlsx")
        assert "error" not in result
        assert "hidden_sheets" in result
        assert any("Hidden" in h for h in result["hidden_sheets"])

    @pytest.mark.asyncio
    async def test_read_returns_number_format(self, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = 0.15
        ws["A1"].number_format = "0.00%"
        wb.save(str(tmp_path / "numfmt.xlsx"))
        wb.close()
        result = await handle(action="read", path="numfmt.xlsx")
        assert "error" not in result
        assert "formatting" in result
        assert "A1" in result["formatting"]
        assert result["formatting"]["A1"]["number_format"] == "0.00%"

    @pytest.mark.asyncio
    async def test_read_plain_file_no_extra_keys(self, tmp_path):
        """A plain file with no formatting should not include extra enrichment keys."""
        result = await handle(
            action="create",
            path="plain.xlsx",
            sheets=[{"name": "Data", "headers": ["A", "B"], "rows": [["1", "2"]]}],
        )
        assert "error" not in result
        result = await handle(action="read", path="plain.xlsx")
        assert "error" not in result
        assert "formulas" not in result
        assert "merged_ranges" not in result

    @pytest.mark.asyncio
    async def test_read_with_cell_range_still_enriched(self, tmp_path):
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "bold"
        ws["A1"].font = Font(bold=True)
        ws["A2"] = "=A1"
        wb.save(str(tmp_path / "range.xlsx"))
        wb.close()
        result = await handle(action="read", path="range.xlsx", cell_range="A1:A2")
        assert "error" not in result
        assert result["rows_read"] == 2


# ---------------------------------------------------------------------------
# template_fill tests
# ---------------------------------------------------------------------------


@_needs_openpyxl
class TestTemplateFill:
    @pytest.mark.asyncio
    async def test_template_fill_basic(self, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Hello {{name}}"
        ws["B1"] = "{{greeting}} World"
        wb.save(str(tmp_path / "tpl.xlsx"))
        wb.close()
        result = await handle(
            action="template_fill",
            path="tpl.xlsx",
            template_data={"name": "Alice", "greeting": "Hi"},
        )
        assert "error" not in result
        assert result["tokens_replaced"] >= 2
        assert result["keys_processed"] == 2
        # Verify replacement
        read = await handle(action="read", path="tpl.xlsx")
        assert "Alice" in read["content"]
        assert "Hi" in read["content"]
        assert "{{name}}" not in read["content"]

    @pytest.mark.asyncio
    async def test_template_fill_across_sheets(self, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1["A1"] = "{{company}}"
        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "{{company}}"
        wb.save(str(tmp_path / "multi.xlsx"))
        wb.close()
        result = await handle(
            action="template_fill",
            path="multi.xlsx",
            template_data={"company": "Acme Corp"},
        )
        assert "error" not in result
        assert result["tokens_replaced"] >= 2

    @pytest.mark.asyncio
    async def test_template_fill_no_data(self):
        result = await handle(action="template_fill", path="test.xlsx")
        assert "error" in result
        assert "template_data is required" in result["error"]

    @pytest.mark.asyncio
    async def test_template_fill_empty_data(self):
        result = await handle(action="template_fill", path="test.xlsx", template_data={})
        assert "error" in result

    @pytest.mark.asyncio
    async def test_template_fill_too_many_keys(self):
        big_data = {f"key{i}": f"val{i}" for i in range(_MAX_EDIT_OPS + 1)}
        result = await handle(action="template_fill", path="test.xlsx", template_data=big_data)
        assert "error" in result
        assert "Too many" in result["error"]

    @pytest.mark.asyncio
    async def test_template_fill_no_matches(self, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "no tokens here"
        wb.save(str(tmp_path / "nomatch.xlsx"))
        wb.close()
        result = await handle(
            action="template_fill",
            path="nomatch.xlsx",
            template_data={"missing": "value"},
        )
        assert "error" not in result
        assert result["tokens_replaced"] == 0

    @pytest.mark.asyncio
    async def test_template_fill_file_not_found(self):
        result = await handle(
            action="template_fill",
            path="missing.xlsx",
            template_data={"key": "val"},
        )
        assert "error" in result

    @pytest.mark.asyncio
    async def test_template_fill_preserves_non_template_cells(self, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "{{name}}"
        ws["B1"] = 42
        ws["C1"] = "=A1"
        wb.save(str(tmp_path / "preserve.xlsx"))
        wb.close()
        result = await handle(
            action="template_fill",
            path="preserve.xlsx",
            template_data={"name": "Bob"},
        )
        assert "error" not in result
        read = await handle(action="read", path="preserve.xlsx")
        import json

        rows = json.loads(read["content"])
        assert rows[0][1] == 42


# ---------------------------------------------------------------------------
# manage_sheets tests
# ---------------------------------------------------------------------------


@_needs_openpyxl
class TestManageSheets:
    @pytest.mark.asyncio
    async def test_rename_sheet(self, tmp_path):
        await _create_test_workbook("manage.xlsx")
        result = await handle(
            action="manage_sheets",
            path="manage.xlsx",
            sheet_operations=[{"op": "rename", "sheet": "Data", "new_name": "Renamed"}],
        )
        assert "error" not in result
        assert result["operations_completed"] == 1
        read = await handle(action="read", path="manage.xlsx")
        assert "Renamed" in read["sheets_available"]

    @pytest.mark.asyncio
    async def test_copy_sheet(self, tmp_path):
        await _create_test_workbook("copy.xlsx")
        result = await handle(
            action="manage_sheets",
            path="copy.xlsx",
            sheet_operations=[{"op": "copy", "sheet": "Data", "new_name": "Data Copy"}],
        )
        assert "error" not in result
        read = await handle(action="read", path="copy.xlsx")
        assert "Data Copy" in read["sheets_available"]

    @pytest.mark.asyncio
    async def test_delete_sheet(self, tmp_path):
        await handle(
            action="create",
            path="del.xlsx",
            sheets=[
                {"name": "Keep", "rows": [["a"]]},
                {"name": "Remove", "rows": [["b"]]},
            ],
        )
        result = await handle(
            action="manage_sheets",
            path="del.xlsx",
            sheet_operations=[{"op": "delete", "sheet": "Remove"}],
        )
        assert "error" not in result
        read = await handle(action="read", path="del.xlsx")
        assert "Remove" not in read["sheets_available"]

    @pytest.mark.asyncio
    async def test_delete_only_sheet_errors(self, tmp_path):
        await _create_test_workbook("solo.xlsx")
        result = await handle(
            action="manage_sheets",
            path="solo.xlsx",
            sheet_operations=[{"op": "delete", "sheet": "Data"}],
        )
        assert "error" in result
        assert "only sheet" in result["error"]

    @pytest.mark.asyncio
    async def test_hide_and_unhide(self, tmp_path):
        await handle(
            action="create",
            path="vis.xlsx",
            sheets=[
                {"name": "Show", "rows": [["a"]]},
                {"name": "Hide", "rows": [["b"]]},
            ],
        )
        result = await handle(
            action="manage_sheets",
            path="vis.xlsx",
            sheet_operations=[{"op": "hide", "sheet": "Hide"}],
        )
        assert "error" not in result
        read = await handle(action="read", path="vis.xlsx")
        assert any("Hide" in h for h in read.get("hidden_sheets", []))

        result = await handle(
            action="manage_sheets",
            path="vis.xlsx",
            sheet_operations=[{"op": "unhide", "sheet": "Hide"}],
        )
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_unknown_op(self, tmp_path):
        await _create_test_workbook("unk.xlsx")
        result = await handle(
            action="manage_sheets",
            path="unk.xlsx",
            sheet_operations=[{"op": "explode", "sheet": "Data"}],
        )
        assert "error" in result
        assert "Unknown sheet operation" in result["error"]

    @pytest.mark.asyncio
    async def test_sheet_not_found(self, tmp_path):
        await _create_test_workbook("nf.xlsx")
        result = await handle(
            action="manage_sheets",
            path="nf.xlsx",
            sheet_operations=[{"op": "rename", "sheet": "Missing", "new_name": "X"}],
        )
        assert "error" in result
        assert "not found" in result["error"]

    @pytest.mark.asyncio
    async def test_no_operations(self):
        result = await handle(action="manage_sheets", path="test.xlsx")
        assert "error" in result
        assert "sheet_operations is required" in result["error"]

    @pytest.mark.asyncio
    async def test_too_many_operations(self):
        ops = [{"op": "rename", "sheet": f"S{i}", "new_name": f"R{i}"} for i in range(_MAX_EDIT_OPS + 1)]
        result = await handle(action="manage_sheets", path="test.xlsx", sheet_operations=ops)
        assert "error" in result
        assert "Too many" in result["error"]

    @pytest.mark.asyncio
    async def test_reorder_sheet(self, tmp_path):
        await handle(
            action="create",
            path="reorder.xlsx",
            sheets=[
                {"name": "First", "rows": [["a"]]},
                {"name": "Second", "rows": [["b"]]},
                {"name": "Third", "rows": [["c"]]},
            ],
        )
        result = await handle(
            action="manage_sheets",
            path="reorder.xlsx",
            sheet_operations=[{"op": "reorder", "sheet": "Third", "position": 1}],
        )
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_multiple_operations(self, tmp_path):
        await handle(
            action="create",
            path="batch.xlsx",
            sheets=[
                {"name": "Alpha", "rows": [["a"]]},
                {"name": "Beta", "rows": [["b"]]},
            ],
        )
        result = await handle(
            action="manage_sheets",
            path="batch.xlsx",
            sheet_operations=[
                {"op": "rename", "sheet": "Alpha", "new_name": "First"},
                {"op": "copy", "sheet": "Beta", "new_name": "Beta Copy"},
            ],
        )
        assert "error" not in result
        assert result["operations_completed"] == 2


# ---------------------------------------------------------------------------
# resize tests
# ---------------------------------------------------------------------------


@_needs_openpyxl
class TestResize:
    @pytest.mark.asyncio
    async def test_resize_column(self, tmp_path):
        await _create_test_workbook("resize.xlsx")
        result = await handle(
            action="resize",
            path="resize.xlsx",
            resize_ops=[{"target": "column", "index": "A", "size": 20.0}],
        )
        assert "error" not in result
        assert result["resized"] == 1
        read = await handle(action="read", path="resize.xlsx")
        assert "column_widths" in read
        assert read["column_widths"]["A"] == 20.0

    @pytest.mark.asyncio
    async def test_resize_row(self, tmp_path):
        await _create_test_workbook("resize_row.xlsx")
        result = await handle(
            action="resize",
            path="resize_row.xlsx",
            resize_ops=[{"target": "row", "index": 1, "size": 30.0}],
        )
        assert "error" not in result
        assert result["resized"] == 1
        read = await handle(action="read", path="resize_row.xlsx")
        assert "row_heights" in read
        assert read["row_heights"][1] == 30.0

    @pytest.mark.asyncio
    async def test_resize_multiple(self, tmp_path):
        await _create_test_workbook("resize_multi.xlsx")
        result = await handle(
            action="resize",
            path="resize_multi.xlsx",
            resize_ops=[
                {"target": "column", "index": "A", "size": 15.0},
                {"target": "column", "index": "B", "size": 25.0},
                {"target": "row", "index": 1, "size": 40.0},
            ],
        )
        assert "error" not in result
        assert result["resized"] == 3

    @pytest.mark.asyncio
    async def test_resize_no_ops(self):
        result = await handle(action="resize", path="test.xlsx")
        assert "error" in result
        assert "resize_ops is required" in result["error"]

    @pytest.mark.asyncio
    async def test_resize_invalid_target(self, tmp_path):
        await _create_test_workbook("inv.xlsx")
        result = await handle(
            action="resize",
            path="inv.xlsx",
            resize_ops=[{"target": "cell", "index": "A", "size": 10}],
        )
        assert "error" in result
        assert "Invalid resize target" in result["error"]

    @pytest.mark.asyncio
    async def test_resize_negative_size(self, tmp_path):
        await _create_test_workbook("neg.xlsx")
        result = await handle(
            action="resize",
            path="neg.xlsx",
            resize_ops=[{"target": "column", "index": "A", "size": -5}],
        )
        assert "error" in result
        assert "non-negative" in result["error"]

    @pytest.mark.asyncio
    async def test_resize_missing_params(self, tmp_path):
        await _create_test_workbook("miss.xlsx")
        result = await handle(
            action="resize",
            path="miss.xlsx",
            resize_ops=[{"target": "column"}],
        )
        assert "error" in result
        assert "'index' and 'size'" in result["error"]


# ---------------------------------------------------------------------------
# insert_delete tests
# ---------------------------------------------------------------------------


@_needs_openpyxl
class TestInsertDelete:
    @pytest.mark.asyncio
    async def test_insert_rows(self, tmp_path):
        await _create_test_workbook("ins.xlsx")
        result = await handle(
            action="insert_delete",
            path="ins.xlsx",
            insert_delete_ops=[{"op": "insert_rows", "index": 2, "count": 3}],
        )
        assert "error" not in result
        assert result["operations_completed"] == 1

    @pytest.mark.asyncio
    async def test_delete_rows(self, tmp_path):
        await handle(
            action="create",
            path="del.xlsx",
            sheets=[{"name": "Data", "rows": [["a"], ["b"], ["c"], ["d"], ["e"]]}],
        )
        result = await handle(
            action="insert_delete",
            path="del.xlsx",
            insert_delete_ops=[{"op": "delete_rows", "index": 2, "count": 2}],
        )
        assert "error" not in result
        read = await handle(action="read", path="del.xlsx")
        assert read["rows_read"] == 3

    @pytest.mark.asyncio
    async def test_insert_cols(self, tmp_path):
        await _create_test_workbook("ins_col.xlsx")
        result = await handle(
            action="insert_delete",
            path="ins_col.xlsx",
            insert_delete_ops=[{"op": "insert_cols", "index": 2, "count": 1}],
        )
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_delete_cols(self, tmp_path):
        await _create_test_workbook("del_col.xlsx")
        result = await handle(
            action="insert_delete",
            path="del_col.xlsx",
            insert_delete_ops=[{"op": "delete_cols", "index": 1, "count": 1}],
        )
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_unknown_op(self, tmp_path):
        await _create_test_workbook("unk.xlsx")
        result = await handle(
            action="insert_delete",
            path="unk.xlsx",
            insert_delete_ops=[{"op": "shuffle", "index": 1}],
        )
        assert "error" in result
        assert "Unknown op" in result["error"]

    @pytest.mark.asyncio
    async def test_no_ops(self):
        result = await handle(action="insert_delete", path="test.xlsx")
        assert "error" in result

    @pytest.mark.asyncio
    async def test_missing_index(self, tmp_path):
        await _create_test_workbook("noidx.xlsx")
        result = await handle(
            action="insert_delete",
            path="noidx.xlsx",
            insert_delete_ops=[{"op": "insert_rows"}],
        )
        assert "error" in result
        assert "'index'" in result["error"]

    @pytest.mark.asyncio
    async def test_multiple_operations(self, tmp_path):
        await handle(
            action="create",
            path="multi.xlsx",
            sheets=[{"name": "Data", "rows": [["a", "b"], ["c", "d"], ["e", "f"]]}],
        )
        result = await handle(
            action="insert_delete",
            path="multi.xlsx",
            insert_delete_ops=[
                {"op": "insert_rows", "index": 2, "count": 1},
                {"op": "insert_cols", "index": 1, "count": 1},
            ],
        )
        assert "error" not in result
        assert result["operations_completed"] == 2

    @pytest.mark.asyncio
    async def test_default_count_is_one(self, tmp_path):
        await handle(
            action="create",
            path="default.xlsx",
            sheets=[{"name": "Data", "rows": [["a"], ["b"], ["c"]]}],
        )
        result = await handle(
            action="insert_delete",
            path="default.xlsx",
            insert_delete_ops=[{"op": "insert_rows", "index": 2}],
        )
        assert "error" not in result
        read = await handle(action="read", path="default.xlsx")
        assert read["rows_read"] == 4  # 3 original + 1 inserted


# ---------------------------------------------------------------------------
# copy_range tests
# ---------------------------------------------------------------------------


@_needs_openpyxl
class TestCopyRange:
    @pytest.mark.asyncio
    async def test_copy_range_basic(self, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Hello"
        ws["A2"] = "World"
        wb.save(str(tmp_path / "cp.xlsx"))
        wb.close()
        result = await handle(
            action="copy_range",
            path="cp.xlsx",
            source_range="A1:A2",
            dest_cell="C1",
        )
        assert "error" not in result
        assert result["cells_copied"] == 2
        read = await handle(action="read", path="cp.xlsx")
        import json

        rows = json.loads(read["content"])
        assert rows[0][2] == "Hello"
        assert rows[1][2] == "World"

    @pytest.mark.asyncio
    async def test_copy_range_values_only(self, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = 10
        ws["A2"] = "=A1*2"
        wb.save(str(tmp_path / "vals.xlsx"))
        wb.close()
        result = await handle(
            action="copy_range",
            path="vals.xlsx",
            source_range="A1:A2",
            dest_cell="C1",
            copy_values_only=True,
        )
        assert "error" not in result
        assert result["values_only"] is True

    @pytest.mark.asyncio
    async def test_copy_range_cross_sheet(self, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Source"
        ws1["A1"] = "data"
        wb.create_sheet("Dest")
        wb.save(str(tmp_path / "cross.xlsx"))
        wb.close()
        result = await handle(
            action="copy_range",
            path="cross.xlsx",
            source_range="A1:A1",
            dest_cell="B1",
            sheet_name="Source",
            dest_sheet="Dest",
        )
        assert "error" not in result

    @pytest.mark.asyncio
    async def test_copy_range_missing_params(self):
        result = await handle(action="copy_range", path="test.xlsx", source_range="A1:A2")
        assert "error" in result
        assert "dest_cell" in result["error"]

    @pytest.mark.asyncio
    async def test_copy_range_missing_source(self):
        result = await handle(action="copy_range", path="test.xlsx", dest_cell="C1")
        assert "error" in result
        assert "source_range" in result["error"]

    @pytest.mark.asyncio
    async def test_copy_range_file_not_found(self):
        result = await handle(
            action="copy_range",
            path="missing.xlsx",
            source_range="A1:A1",
            dest_cell="B1",
        )
        assert "error" in result

    @pytest.mark.asyncio
    async def test_copy_range_with_formatting(self, tmp_path):
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "bold"
        ws["A1"].font = Font(bold=True)
        wb.save(str(tmp_path / "fmtcp.xlsx"))
        wb.close()
        result = await handle(
            action="copy_range",
            path="fmtcp.xlsx",
            source_range="A1:A1",
            dest_cell="C1",
            copy_values_only=False,
        )
        assert "error" not in result
        # Verify formatting was copied
        wb2 = openpyxl.load_workbook(str(tmp_path / "fmtcp.xlsx"))
        assert wb2.active["C1"].font.bold is True
        wb2.close()

    @pytest.mark.asyncio
    async def test_copy_range_dest_sheet_not_found(self, tmp_path):
        await _create_test_workbook("dst.xlsx")
        result = await handle(
            action="copy_range",
            path="dst.xlsx",
            source_range="A1:A1",
            dest_cell="B1",
            dest_sheet="Nonexistent",
        )
        assert "error" in result
        assert "not found" in result["error"]


# ---------------------------------------------------------------------------
# New action definition tests
# ---------------------------------------------------------------------------


class TestNewActionDefinitions:
    def test_new_actions_in_definition(self):
        actions = DEFINITION["parameters"]["properties"]["action"]["enum"]
        assert "template_fill" in actions
        assert "manage_sheets" in actions
        assert "resize" in actions
        assert "insert_delete" in actions
        assert "copy_range" in actions

    def test_template_data_param(self):
        props = DEFINITION["parameters"]["properties"]
        assert "template_data" in props
        assert props["template_data"]["type"] == "object"

    def test_sheet_operations_param(self):
        props = DEFINITION["parameters"]["properties"]
        assert "sheet_operations" in props
        assert props["sheet_operations"]["type"] == "array"

    def test_resize_ops_param(self):
        props = DEFINITION["parameters"]["properties"]
        assert "resize_ops" in props

    def test_insert_delete_ops_param(self):
        props = DEFINITION["parameters"]["properties"]
        assert "insert_delete_ops" in props

    def test_copy_range_params(self):
        props = DEFINITION["parameters"]["properties"]
        assert "source_range" in props
        assert "dest_cell" in props
        assert "copy_values_only" in props
        assert "dest_sheet" in props
