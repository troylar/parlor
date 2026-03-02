"""XLSX (Excel spreadsheet) create/read/edit tool.

Backends:
- COM (Windows + Office + pywin32): full Office object model
- Library (openpyxl): cross-platform XML manipulation

Install: ``pip install anteroom[office]`` or ``pip install anteroom[office-com]``
"""

from __future__ import annotations

import json
import os
import re
import sys
from typing import Any

from .security import validate_path

_BACKEND: str | None = None
_com_mod: Any = None

if sys.platform == "win32":
    try:
        from . import office_com as _com_mod_import

        if _com_mod_import.COM_AVAILABLE:
            _com_mod = _com_mod_import
            _BACKEND = "com"
    except ImportError:
        pass

if _BACKEND is None:
    try:
        import openpyxl  # noqa: F401

        _BACKEND = "lib"
    except ImportError:
        pass

AVAILABLE = _BACKEND is not None

_MAX_OUTPUT = 100_000
_MAX_CONTENT_BLOCKS = 200
_MAX_ROWS = 10_000
_MAX_EDIT_OPS = 500
_MAX_EXCEL_COLUMNS = 16_384  # Excel column limit (XFD)
_MAX_EXCEL_ROWS = 1_048_576  # Excel row limit

_VALID_CHART_TYPES = ("column", "bar", "line", "pie", "scatter", "area")
_VALID_CF_OPERATORS = (
    "greaterThan",
    "lessThan",
    "equal",
    "notEqual",
    "between",
    "greaterThanOrEqual",
    "lessThanOrEqual",
)

_ALL_ACTIONS = [
    "create",
    "read",
    "edit",
    "format_cells",
    "merge_cells",
    "freeze_panes",
    "auto_filter",
    "print_area",
    "named_ranges",
    "data_validation",
    "conditional_format",
    "comments",
    "hyperlinks",
    "images",
    "protect",
    "group_rows_cols",
    "print_settings",
    "charts",
    "export_pdf",
    "sort",
    "pivot_tables",
    "sparklines",
    "slicers",
    "template_fill",
    "manage_sheets",
    "resize",
    "insert_delete",
    "copy_range",
]

_working_dir_override: str | None = None

DEFINITION: dict[str, Any] = {
    "name": "xlsx",
    "description": (
        "Create, read, edit, and manipulate Excel spreadsheets (.xlsx). "
        "Read returns values, formulas, merged ranges, cell formatting, column widths, "
        "data validations, conditional formatting rules, named ranges, and freeze pane state. "
        "Supports formatting, charts, pivot tables, template fill ({{key}} replacement), "
        "sheet management (rename/delete/copy/reorder/hide), column/row resize, "
        "bulk insert/delete rows/columns, range copy, and more."
    ),
    "parameters": {
        "type": "object",
        "properties": {
            "action": {
                "type": "string",
                "enum": _ALL_ACTIONS,
                "description": "Operation to perform.",
            },
            "path": {
                "type": "string",
                "description": "File path (relative to working directory or absolute).",
            },
            "sheet_name": {
                "type": "string",
                "description": "Target sheet name. Defaults to active sheet.",
            },
            "sheets": {
                "type": "array",
                "description": "Sheets for create. Each: {name, rows?, headers?}.",
                "items": {"type": "object"},
            },
            "cell_range": {
                "type": "string",
                "description": "Cell range, e.g. 'A1:C10'.",
            },
            "updates": {
                "type": "array",
                "description": "Cell updates for edit: [{cell, value}].",
                "items": {"type": "object"},
            },
            "append_rows": {
                "type": "array",
                "description": "Rows to append: [[value, ...]].",
                "items": {"type": "array", "items": {}},
            },
            "add_sheets": {
                "type": "array",
                "description": "New sheets for edit: [{name, rows?}].",
                "items": {"type": "object"},
            },
            "format": {
                "type": "object",
                "description": (
                    "Cell formatting for format_cells. Keys: font (name, size, bold, italic, color), "
                    "fill (color, pattern), border (style, color), number_format (str), "
                    "alignment (horizontal, vertical, wrap_text)."
                ),
            },
            "merge": {
                "type": "boolean",
                "description": "True to merge, false to unmerge (merge_cells action).",
            },
            "row": {
                "type": "integer",
                "description": "Row number for freeze_panes (freeze above this row).",
            },
            "column": {
                "type": "integer",
                "description": "Column number for freeze_panes (freeze left of this column).",
            },
            "range_name": {
                "type": "string",
                "description": "Named range name for named_ranges action.",
            },
            "operation": {
                "type": "string",
                "enum": ["add", "list", "delete", "read", "set", "enable", "disable", "group", "ungroup"],
                "description": "Sub-operation for actions that support multiple modes.",
            },
            "validation": {
                "type": "object",
                "description": (
                    "Validation rule for data_validation: "
                    "{type: 'list'|'whole'|'decimal'|'date'|'textLength', "
                    "formula1, formula2?, operator?, allow_blank?, error_message?}."
                ),
            },
            "rule": {
                "type": "object",
                "description": (
                    "Conditional format rule: {type: 'cellIs'|'colorScale'|'dataBar'|'expression', "
                    "operator?, formula?, format?}."
                ),
            },
            "comment_text": {
                "type": "string",
                "description": "Comment text for comments action.",
            },
            "author": {
                "type": "string",
                "description": "Comment author name.",
            },
            "url": {
                "type": "string",
                "description": "URL for hyperlinks action.",
            },
            "display_text": {
                "type": "string",
                "description": "Display text for hyperlinks.",
            },
            "image_path": {
                "type": "string",
                "description": "Image file path for images action.",
            },
            "anchor_cell": {
                "type": "string",
                "description": "Cell to anchor image/chart to, e.g. 'E2'.",
            },
            "password": {
                "type": "string",
                "description": "Password for protect action.",
            },
            "protect_options": {
                "type": "object",
                "description": (
                    "Protection options: {sheet?, format_cells?, insert_rows?, delete_rows?, sort?, auto_filter?}."
                ),
            },
            "start": {
                "type": "integer",
                "description": "Start row/column index for group_rows_cols.",
            },
            "end": {
                "type": "integer",
                "description": "End row/column index for group_rows_cols.",
            },
            "axis": {
                "type": "string",
                "enum": ["rows", "columns"],
                "description": "Axis for group_rows_cols.",
            },
            "page_setup": {
                "type": "object",
                "description": (
                    "Print settings: {orientation: 'portrait'|'landscape', paper_size?, "
                    "fit_to_width?, fit_to_height?, header?, footer?, margins?}."
                ),
            },
            "chart_type": {
                "type": "string",
                "description": "Chart type: 'bar', 'line', 'pie', 'scatter', 'area', 'column'.",
            },
            "chart_title": {
                "type": "string",
                "description": "Chart title text.",
            },
            "data_range": {
                "type": "string",
                "description": "Data range for charts/pivot_tables, e.g. 'A1:D10'.",
            },
            "sort_column": {
                "type": "string",
                "description": "Column letter or cell ref for sort, e.g. 'B'.",
            },
            "ascending": {
                "type": "boolean",
                "description": "Sort ascending (true) or descending (false). Default true.",
            },
            "output_path": {
                "type": "string",
                "description": "Output file path for export_pdf.",
            },
            "template_data": {
                "type": "object",
                "description": (
                    "Key-value pairs for template_fill action. "
                    "Each {{key}} in cells is replaced with the corresponding value."
                ),
            },
            "sheet_operations": {
                "type": "array",
                "description": (
                    "Operations for manage_sheets: [{op: 'rename'|'delete'|'copy'|'reorder'|'hide'|'unhide', "
                    "sheet: str, new_name?: str, position?: int, target_name?: str}]."
                ),
                "items": {"type": "object"},
            },
            "resize_ops": {
                "type": "array",
                "description": (
                    "Resize operations: [{target: 'column'|'row', index: str|int, size: float}]. "
                    "Column index is letter (e.g. 'A'), row index is number. "
                    "Size is width for columns (character units) or height for rows (points)."
                ),
                "items": {"type": "object"},
            },
            "insert_delete_ops": {
                "type": "array",
                "description": (
                    "Insert/delete operations: [{op: 'insert_rows'|'delete_rows'|'insert_cols'|'delete_cols', "
                    "index: int, count?: int}]. index is 1-based."
                ),
                "items": {"type": "object"},
            },
            "source_range": {
                "type": "string",
                "description": "Source cell range for copy_range, e.g. 'A1:C10'.",
            },
            "dest_cell": {
                "type": "string",
                "description": "Destination top-left cell for copy_range, e.g. 'E1'.",
            },
            "copy_values_only": {
                "type": "boolean",
                "description": "If true, copy only values (not formulas). Default false.",
            },
            "dest_sheet": {
                "type": "string",
                "description": "Destination sheet name for copy_range (defaults to source sheet).",
            },
        },
        "required": ["action", "path"],
    },
}


def set_working_dir(d: str) -> None:
    """Override the working directory used for path resolution.

    When set, this value is used instead of os.getcwd(). Pass an empty
    string to clear the override and revert to os.getcwd().
    """
    global _working_dir_override
    _working_dir_override = d if d else None


def _open_workbook_lib(resolved: str, display_path: str, read_only: bool = False) -> tuple[Any, str | None]:
    """Open a workbook with openpyxl, returning (wb, error)."""
    import openpyxl as _openpyxl

    if not os.path.isfile(resolved):
        return None, f"File not found: {display_path}"
    try:
        wb = _openpyxl.load_workbook(resolved, read_only=read_only, data_only=read_only)
        return wb, None
    except Exception as exc:
        return None, f"Unable to read XLSX file: {display_path} ({type(exc).__name__}: {exc})"


def _get_sheet_lib(wb: Any, sheet_name: str | None) -> tuple[Any, str | None]:
    """Get a worksheet by name or active, returning (ws, error)."""
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            return None, f"Sheet '{sheet_name}' not found. Available: {list(wb.sheetnames)}"
        return wb[sheet_name], None
    ws = wb.active
    if ws is None:
        return None, "No active sheet found"
    return ws, None


async def handle(action: str, path: str, **kwargs: Any) -> dict[str, Any]:
    if not AVAILABLE:
        return {"error": "No xlsx backend available. Install with: pip install anteroom[office]"}

    working_dir = _working_dir_override or os.getcwd()
    resolved, error = validate_path(path, working_dir)
    if error:
        return {"error": error}

    if _BACKEND == "com":
        return await _dispatch_com(action, resolved, path, working_dir=working_dir, **kwargs)

    # Library backend dispatch
    _lib_dispatch: dict[str, Any] = {
        "create": _create_lib,
        "read": _read_lib,
        "edit": _edit_lib,
        "format_cells": _format_cells_lib,
        "merge_cells": _merge_cells_lib,
        "freeze_panes": _freeze_panes_lib,
        "auto_filter": _auto_filter_lib,
        "print_area": _print_area_lib,
        "named_ranges": _named_ranges_lib,
        "data_validation": _data_validation_lib,
        "conditional_format": _conditional_format_lib,
        "comments": _comments_lib,
        "hyperlinks": _hyperlinks_lib,
        "images": _images_lib,
        "protect": _protect_lib,
        "group_rows_cols": _group_rows_cols_lib,
        "print_settings": _print_settings_lib,
        "charts": _charts_lib,
        "export_pdf": _export_pdf_lib,
        "sort": _sort_lib,
        "pivot_tables": _pivot_tables_lib,
        "sparklines": _sparklines_lib,
        "slicers": _slicers_lib,
        "template_fill": _template_fill_lib,
        "manage_sheets": _manage_sheets_lib,
        "resize": _resize_lib,
        "insert_delete": _insert_delete_lib,
        "copy_range": _copy_range_lib,
    }

    handler = _lib_dispatch.get(action)
    if handler is None:
        return {"error": f"Unknown action: {action}. Available: {', '.join(_ALL_ACTIONS)}"}
    return dict(handler(resolved, path, working_dir=working_dir, **kwargs))


# ---------------------------------------------------------------------------
# COM backend
# ---------------------------------------------------------------------------

_COM_ACTIONS = [
    "create",
    "read",
    "edit",
    "format_cells",
    "merge_cells",
    "freeze_panes",
    "auto_filter",
    "print_area",
    "named_ranges",
    "data_validation",
    "conditional_format",
    "comments",
    "hyperlinks",
    "images",
    "protect",
    "group_rows_cols",
    "print_settings",
    "charts",
    "export_pdf",
    "sort",
    "pivot_tables",
    "sparklines",
    "slicers",
    "template_fill",
    "manage_sheets",
    "resize",
    "insert_delete",
    "copy_range",
]


async def _dispatch_com(
    action: str,
    resolved: str,
    display_path: str,
    *,
    working_dir: str,
    **kwargs: Any,
) -> dict[str, Any]:
    manager = _com_mod.get_manager()
    kwargs["working_dir"] = working_dir

    _com_dispatch: dict[str, Any] = {
        "create": _create_com,
        "read": _read_com,
        "edit": _edit_com,
        "format_cells": _format_cells_com,
        "merge_cells": _merge_cells_com,
        "freeze_panes": _freeze_panes_com,
        "auto_filter": _auto_filter_com,
        "print_area": _print_area_com,
        "named_ranges": _named_ranges_com,
        "data_validation": _data_validation_com,
        "conditional_format": _conditional_format_com,
        "comments": _comments_com,
        "hyperlinks": _hyperlinks_com,
        "images": _images_com,
        "protect": _protect_com,
        "group_rows_cols": _group_rows_cols_com,
        "print_settings": _print_settings_com,
        "charts": _charts_com,
        "export_pdf": _export_pdf_com,
        "sort": _sort_com,
        "pivot_tables": _pivot_tables_com,
        "sparklines": _sparklines_com,
        "slicers": _slicers_com,
        "template_fill": _template_fill_com,
        "manage_sheets": _manage_sheets_com,
        "resize": _resize_com,
        "insert_delete": _insert_delete_com,
        "copy_range": _copy_range_com,
    }

    handler = _com_dispatch.get(action)
    if handler is None:
        return {"error": f"Unknown action: {action}. Available: {', '.join(_ALL_ACTIONS)}"}
    try:
        return dict(await manager.run_com(handler, manager, resolved, display_path, **kwargs))
    except Exception as exc:
        return {"error": f"COM {action} failed on {display_path}: {type(exc).__name__}: {exc}"}


def _open_workbook_com(
    manager: Any,
    resolved: str,
    display_path: str,
    read_only: bool = False,
) -> tuple[Any, Any, str | None]:
    """Open workbook via COM. Returns (excel, wb, error)."""
    if not os.path.isfile(resolved):
        return None, None, f"File not found: {display_path}"
    excel = manager.get_app("Excel.Application")
    try:
        wb = excel.Workbooks.Open(os.path.abspath(resolved), ReadOnly=read_only)
        return excel, wb, None
    except Exception as exc:
        return None, None, f"Unable to read XLSX file: {display_path} ({type(exc).__name__}: {exc})"


def _get_sheet_com(wb: Any, sheet_name: str | None) -> tuple[Any, str | None]:
    """Get worksheet by name or active via COM."""
    if sheet_name:
        names = [wb.Worksheets(i).Name for i in range(1, wb.Worksheets.Count + 1)]
        if sheet_name not in names:
            return None, f"Sheet '{sheet_name}' not found. Available: {names}"
        return wb.Worksheets(sheet_name), None
    return wb.ActiveSheet, None


# --- create/read/edit (unchanged from #588) ---


def _create_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    sheets: list[dict[str, Any]] = kwargs.get("sheets") or []
    if not sheets:
        return {"error": "sheets is required for create action"}
    if len(sheets) > _MAX_CONTENT_BLOCKS:
        return {"error": f"Too many sheets (max {_MAX_CONTENT_BLOCKS})"}

    excel = manager.get_app("Excel.Application")
    wb = excel.Workbooks.Add()

    try:
        while wb.Worksheets.Count > 1:
            wb.Worksheets(wb.Worksheets.Count).Delete()

        total_rows = 0
        for idx, sheet_def in enumerate(sheets):
            name = str(sheet_def.get("name", "Sheet"))
            if idx == 0:
                ws = wb.Worksheets(1)
                ws.Name = name
            else:
                ws = wb.Worksheets.Add(After=wb.Worksheets(wb.Worksheets.Count))
                ws.Name = name

            headers: list[str] = sheet_def.get("headers") or []
            rows: list[list[Any]] = sheet_def.get("rows") or []

            row_num = 1
            if headers:
                for col_num, val in enumerate(headers, 1):
                    ws.Cells(row_num, col_num).Value = _sanitize_cell_value(val)
                row_num += 1
                total_rows += 1

            for row in rows:
                if total_rows >= _MAX_ROWS:
                    break
                for col_num, val in enumerate(row, 1):
                    ws.Cells(row_num, col_num).Value = _sanitize_cell_value(val)
                row_num += 1
                total_rows += 1

        if total_rows >= _MAX_ROWS:
            return {"error": f"Too many rows (max {_MAX_ROWS})"}

        os.makedirs(os.path.dirname(resolved) or ".", exist_ok=True)
        wb.SaveAs(os.path.abspath(resolved), FileFormat=51)
    finally:
        wb.Close(SaveChanges=False)

    return {
        "result": f"Created {display_path}",
        "path": display_path,
        "sheets_created": len(sheets),
        "total_rows": total_rows,
    }


def _read_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    _, wb, err = _open_workbook_com(manager, resolved, display_path, read_only=True)
    if err:
        return {"error": err}

    try:
        sheet_name: str | None = kwargs.get("sheet_name")
        cell_range: str | None = kwargs.get("cell_range")

        available = [wb.Worksheets(i).Name for i in range(1, wb.Worksheets.Count + 1)]
        ws, err = _get_sheet_com(wb, sheet_name)
        if err:
            return {"error": err}

        output_rows: list[list[Any]] = []
        formulas: dict[str, str] = {}
        formatting: dict[str, dict[str, Any]] = {}

        if cell_range:
            rng = ws.Range(cell_range)
            for r in range(1, rng.Rows.Count + 1):
                row_data = []
                for c in range(1, rng.Columns.Count + 1):
                    cell = rng.Cells(r, c)
                    row_data.append(cell.Value)
                    # Collect formula
                    if cell.HasFormula:
                        formulas[cell.Address.replace("$", "")] = cell.Formula
                    # Collect formatting
                    fmt_info = _collect_com_cell_format(cell)
                    if fmt_info:
                        formatting[cell.Address.replace("$", "")] = fmt_info
                output_rows.append(row_data)
                if len(output_rows) >= _MAX_ROWS:
                    break
        else:
            used = ws.UsedRange
            if used is not None:
                for r in range(1, used.Rows.Count + 1):
                    row_data = []
                    for c in range(1, used.Columns.Count + 1):
                        cell = used.Cells(r, c)
                        row_data.append(cell.Value)
                        if cell.HasFormula:
                            formulas[cell.Address.replace("$", "")] = cell.Formula
                        fmt_info = _collect_com_cell_format(cell)
                        if fmt_info:
                            formatting[cell.Address.replace("$", "")] = fmt_info
                    output_rows.append(row_data)
                    if len(output_rows) >= _MAX_ROWS:
                        break

        # Collect merged ranges
        merged_ranges: list[str] = []
        try:
            for area in ws.Cells.MergeCells if ws.Cells.MergeCells else []:
                merged_ranges.append(str(area))
        except Exception:
            try:
                used = ws.UsedRange
                if used:
                    for r in range(1, min(used.Rows.Count + 1, _MAX_ROWS)):
                        for c in range(1, used.Columns.Count + 1):
                            cell = used.Cells(r, c)
                            if cell.MergeCells:
                                ma = cell.MergeArea.Address.replace("$", "")
                                if ma not in merged_ranges:
                                    merged_ranges.append(ma)
            except Exception:
                pass

        # Collect freeze pane
        freeze_pane: str | None = None
        try:
            win = wb.Application.ActiveWindow
            if win.FreezePanes:
                freeze_pane = f"Row {win.SplitRow + 1}, Col {win.SplitColumn + 1}"
        except Exception:
            pass

        # Collect named ranges
        named_ranges: list[dict[str, str]] = []
        try:
            for i in range(1, wb.Names.Count + 1):
                n = wb.Names(i)
                named_ranges.append({"name": n.Name, "refers_to": n.RefersTo})
        except Exception:
            pass

        # Hidden sheets
        hidden_sheets: list[str] = []
        try:
            for i in range(1, wb.Worksheets.Count + 1):
                s = wb.Worksheets(i)
                if s.Visible != -1:  # xlSheetVisible = -1
                    hidden_sheets.append(f"{s.Name} (hidden)")
        except Exception:
            pass

        sheet_title = ws.Name
    finally:
        wb.Close(SaveChanges=False)

    content = json.dumps(output_rows, ensure_ascii=False, default=str)
    if len(content) > _MAX_OUTPUT:
        content = content[:_MAX_OUTPUT] + "\n... (truncated)"

    result: dict[str, Any] = {
        "content": content,
        "sheet": sheet_title,
        "sheets_available": available,
        "rows_read": len(output_rows),
    }

    if formulas:
        result["formulas"] = formulas
    if formatting:
        result["formatting"] = formatting
    if merged_ranges:
        result["merged_ranges"] = merged_ranges
    if freeze_pane:
        result["freeze_pane"] = freeze_pane
    if named_ranges:
        result["named_ranges"] = named_ranges
    if hidden_sheets:
        result["hidden_sheets"] = hidden_sheets

    return result


def _collect_com_cell_format(cell: Any) -> dict[str, Any]:
    """Extract non-default formatting from a COM cell object."""
    fmt: dict[str, Any] = {}
    try:
        f = cell.Font
        if f.Bold:
            fmt["bold"] = True
        if f.Italic:
            fmt["italic"] = True
        if f.Underline and f.Underline != -4142:  # xlUnderlineStyleNone
            fmt["underline"] = True
        if f.Name and f.Name != "Calibri":
            fmt["font"] = f.Name
        if f.Size and f.Size != 11:
            fmt["size"] = f.Size
    except Exception:
        pass
    try:
        if cell.Interior.ColorIndex not in (None, -4142, 0):  # xlNone
            color_int = cell.Interior.Color
            r = color_int & 0xFF
            g = (color_int >> 8) & 0xFF
            b = (color_int >> 16) & 0xFF
            fmt["fill"] = f"#{r:02X}{g:02X}{b:02X}"
    except Exception:
        pass
    try:
        nf = cell.NumberFormat
        if nf and nf != "General":
            fmt["number_format"] = nf
    except Exception:
        pass
    return fmt


def _edit_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    # NOTE: COM and lib backends may report slightly different counts for
    # cells_updated when a cell reference is invalid or skipped. COM silently
    # applies Range assignments while lib uses dict-style ws[cell] access.
    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    sheet_name: str | None = kwargs.get("sheet_name")
    updates: list[dict[str, Any]] = kwargs.get("updates") or []
    append_rows: list[list[Any]] = kwargs.get("append_rows") or []
    add_sheets: list[dict[str, Any]] = kwargs.get("add_sheets") or []

    try:
        if not updates and not append_rows and not add_sheets:
            return {"error": "Provide 'updates', 'append_rows', and/or 'add_sheets' for edit action"}

        if append_rows and len(append_rows) > _MAX_ROWS:
            return {"error": f"Too many rows to append (max {_MAX_ROWS})"}

        cells_updated = 0
        rows_appended = 0
        sheets_added = 0

        if updates or append_rows:
            ws, err = _get_sheet_com(wb, sheet_name)
            if err:
                return {"error": err}

            for upd in updates:
                cell = upd.get("cell", "")
                value = upd.get("value")
                if not cell:
                    continue
                ws.Range(cell).Value = _sanitize_cell_value(value)
                cells_updated += 1

            if append_rows:
                used = ws.UsedRange
                next_row = used.Row + used.Rows.Count if used is not None else 1
                for row in append_rows:
                    for col_num, val in enumerate(row, 1):
                        ws.Cells(next_row, col_num).Value = _sanitize_cell_value(val)
                    next_row += 1
                    rows_appended += 1

        for sheet_def in add_sheets:
            name = str(sheet_def.get("name", "Sheet"))
            ws_new = wb.Worksheets.Add(After=wb.Worksheets(wb.Worksheets.Count))
            ws_new.Name = name
            rows = sheet_def.get("rows") or []
            for r_idx, row in enumerate(rows[:_MAX_ROWS], 1):
                for c_idx, val in enumerate(row, 1):
                    ws_new.Cells(r_idx, c_idx).Value = _sanitize_cell_value(val)
            sheets_added += 1

        wb.Save()
    finally:
        wb.Close(SaveChanges=False)

    return {
        "result": f"Edited {display_path}",
        "path": display_path,
        "cells_updated": cells_updated,
        "rows_appended": rows_appended,
        "sheets_added": sheets_added,
    }


# --- format_cells ---


def _format_cells_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        ws, err = _get_sheet_com(wb, kwargs.get("sheet_name"))
        if err:
            return {"error": err}

        cell_range = kwargs.get("cell_range")
        if not cell_range:
            return {"error": "cell_range is required for format_cells"}

        fmt = kwargs.get("format") or {}
        rng = ws.Range(cell_range)

        font = fmt.get("font") or {}
        if font.get("name"):
            rng.Font.Name = font["name"]
        if font.get("size"):
            rng.Font.Size = font["size"]
        if font.get("bold") is not None:
            rng.Font.Bold = font["bold"]
        if font.get("italic") is not None:
            rng.Font.Italic = font["italic"]
        if font.get("color"):
            rng.Font.Color = _parse_color_int(font["color"])

        fill = fmt.get("fill") or {}
        if fill.get("color"):
            rng.Interior.Color = _parse_color_int(fill["color"])

        if fmt.get("number_format"):
            rng.NumberFormat = fmt["number_format"]

        alignment = fmt.get("alignment") or {}
        if alignment.get("horizontal"):
            h_map = {"left": -4131, "center": -4108, "right": -4152}
            rng.HorizontalAlignment = h_map.get(alignment["horizontal"], -4131)
        if alignment.get("vertical"):
            v_map = {"top": -4160, "center": -4108, "bottom": -4107}
            rng.VerticalAlignment = v_map.get(alignment["vertical"], -4107)
        if alignment.get("wrap_text") is not None:
            rng.WrapText = alignment["wrap_text"]

        wb.Save()
    finally:
        wb.Close(SaveChanges=False)

    return {"result": f"Formatted {cell_range} in {display_path}", "path": display_path}


def _format_cells_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    from openpyxl.styles import Alignment, Font, PatternFill

    wb, err = _open_workbook_lib(resolved, display_path)
    if err:
        return {"error": err}

    ws, err = _get_sheet_lib(wb, kwargs.get("sheet_name"))
    if err:
        wb.close()
        return {"error": err}

    cell_range = kwargs.get("cell_range")
    if not cell_range:
        wb.close()
        return {"error": "cell_range is required for format_cells"}

    fmt = kwargs.get("format") or {}

    try:
        font_opts = fmt.get("font") or {}
        font_kwargs: dict[str, Any] = {}
        if font_opts.get("name"):
            font_kwargs["name"] = font_opts["name"]
        if font_opts.get("size"):
            font_kwargs["size"] = font_opts["size"]
        if font_opts.get("bold") is not None:
            font_kwargs["bold"] = font_opts["bold"]
        if font_opts.get("italic") is not None:
            font_kwargs["italic"] = font_opts["italic"]
        if font_opts.get("color"):
            font_kwargs["color"] = _normalize_hex_color(font_opts["color"])
        font = Font(**font_kwargs) if font_kwargs else None

        fill_opts = fmt.get("fill") or {}
        fill = None
        if fill_opts.get("color"):
            fill = PatternFill(start_color=_normalize_hex_color(fill_opts["color"]), fill_type="solid")

        align_opts = fmt.get("alignment") or {}
        align = None
        if align_opts:
            align = Alignment(
                horizontal=align_opts.get("horizontal"),
                vertical=align_opts.get("vertical"),
                wrap_text=align_opts.get("wrap_text"),
            )

        number_format = fmt.get("number_format")

        cells = ws[cell_range]
        # Single cell returns a Cell object, not a tuple of tuples
        if not isinstance(cells, tuple):
            cells = ((cells,),)
        for row in cells:
            for cell in row:
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill
                if align:
                    cell.alignment = align
                if number_format:
                    cell.number_format = number_format

        wb.save(resolved)
    finally:
        wb.close()
    return {"result": f"Formatted {cell_range} in {display_path}", "path": display_path}


# --- merge_cells ---


def _merge_cells_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        ws, err = _get_sheet_com(wb, kwargs.get("sheet_name"))
        if err:
            return {"error": err}

        cell_range = kwargs.get("cell_range")
        if not cell_range:
            return {"error": "cell_range is required for merge_cells"}

        merge = kwargs.get("merge", True)
        if merge:
            ws.Range(cell_range).Merge()
        else:
            ws.Range(cell_range).UnMerge()

        wb.Save()
    finally:
        wb.Close(SaveChanges=False)

    action_word = "Merged" if kwargs.get("merge", True) else "Unmerged"
    return {"result": f"{action_word} {cell_range} in {display_path}", "path": display_path}


def _merge_cells_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    wb, err = _open_workbook_lib(resolved, display_path)
    if err:
        return {"error": err}

    ws, err = _get_sheet_lib(wb, kwargs.get("sheet_name"))
    if err:
        wb.close()
        return {"error": err}

    cell_range = kwargs.get("cell_range")
    if not cell_range:
        wb.close()
        return {"error": "cell_range is required for merge_cells"}

    merge = kwargs.get("merge", True)
    if merge:
        ws.merge_cells(cell_range)
    else:
        ws.unmerge_cells(cell_range)

    wb.save(resolved)
    wb.close()
    action_word = "Merged" if merge else "Unmerged"
    return {"result": f"{action_word} {cell_range} in {display_path}", "path": display_path}


# --- freeze_panes ---


def _freeze_panes_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        ws, err = _get_sheet_com(wb, kwargs.get("sheet_name"))
        if err:
            return {"error": err}

        row = kwargs.get("row", 2)
        col = kwargs.get("column", 1)
        ws.Activate()
        wb.Application.ActiveWindow.SplitRow = row - 1
        wb.Application.ActiveWindow.SplitColumn = col - 1
        wb.Application.ActiveWindow.FreezePanes = True
        wb.Save()
    finally:
        wb.Close(SaveChanges=False)

    return {"result": f"Froze panes at row {row}, column {col} in {display_path}", "path": display_path}


def _freeze_panes_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    from openpyxl.utils import get_column_letter

    wb, err = _open_workbook_lib(resolved, display_path)
    if err:
        return {"error": err}

    ws, err = _get_sheet_lib(wb, kwargs.get("sheet_name"))
    if err:
        wb.close()
        return {"error": err}

    row = kwargs.get("row", 2)
    col = kwargs.get("column", 1)
    ws.freeze_panes = f"{get_column_letter(col)}{row}"
    wb.save(resolved)
    wb.close()
    return {"result": f"Froze panes at row {row}, column {col} in {display_path}", "path": display_path}


# --- auto_filter ---


def _auto_filter_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        ws, err = _get_sheet_com(wb, kwargs.get("sheet_name"))
        if err:
            return {"error": err}

        cell_range = kwargs.get("cell_range")
        op = kwargs.get("operation", "enable")

        if op == "disable":
            if ws.AutoFilterMode:
                ws.AutoFilterMode = False
        else:
            rng = ws.Range(cell_range) if cell_range else ws.UsedRange
            rng.AutoFilter()

        wb.Save()
    finally:
        wb.Close(SaveChanges=False)

    return {"result": f"Auto-filter updated in {display_path}", "path": display_path}


def _auto_filter_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    wb, err = _open_workbook_lib(resolved, display_path)
    if err:
        return {"error": err}

    ws, err = _get_sheet_lib(wb, kwargs.get("sheet_name"))
    if err:
        wb.close()
        return {"error": err}

    cell_range = kwargs.get("cell_range")
    op = kwargs.get("operation", "enable")

    if op == "disable":
        ws.auto_filter.ref = None
    else:
        if cell_range:
            ws.auto_filter.ref = cell_range
        elif ws.dimensions:
            ws.auto_filter.ref = ws.dimensions

    wb.save(resolved)
    wb.close()
    return {"result": f"Auto-filter updated in {display_path}", "path": display_path}


# --- print_area ---


def _print_area_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        ws, err = _get_sheet_com(wb, kwargs.get("sheet_name"))
        if err:
            return {"error": err}

        cell_range = kwargs.get("cell_range")
        if not cell_range:
            return {"error": "cell_range is required for print_area"}

        ws.PageSetup.PrintArea = cell_range
        wb.Save()
    finally:
        wb.Close(SaveChanges=False)

    return {"result": f"Set print area to {cell_range} in {display_path}", "path": display_path}


def _print_area_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    wb, err = _open_workbook_lib(resolved, display_path)
    if err:
        return {"error": err}

    ws, err = _get_sheet_lib(wb, kwargs.get("sheet_name"))
    if err:
        wb.close()
        return {"error": err}

    cell_range = kwargs.get("cell_range")
    if not cell_range:
        wb.close()
        return {"error": "cell_range is required for print_area"}

    ws.print_area = cell_range
    wb.save(resolved)
    wb.close()
    return {"result": f"Set print area to {cell_range} in {display_path}", "path": display_path}


# --- named_ranges ---


def _named_ranges_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        op = kwargs.get("operation", "list")
        range_name = kwargs.get("range_name")

        if op == "list":
            names = []
            for i in range(1, wb.Names.Count + 1):
                n = wb.Names(i)
                names.append({"name": n.Name, "refers_to": n.RefersTo})
            return {"result": "Listed named ranges", "names": names}

        if op == "add":
            cell_range = kwargs.get("cell_range")
            sheet_name = kwargs.get("sheet_name")
            if not range_name or not cell_range:
                return {"error": "range_name and cell_range required for add"}
            if sheet_name:
                ref = f"='{sheet_name}'!{cell_range}"
            else:
                ref = f"='{wb.ActiveSheet.Name}'!{cell_range}"
            wb.Names.Add(Name=range_name, RefersTo=ref)
            wb.Save()
            return {"result": f"Added named range '{range_name}'", "path": display_path}

        if op == "delete":
            if not range_name:
                return {"error": "range_name required for delete"}
            wb.Names(range_name).Delete()
            wb.Save()
            return {"result": f"Deleted named range '{range_name}'", "path": display_path}
    finally:
        try:
            wb.Close(SaveChanges=False)
        except Exception:  # COM cleanup; must not mask the primary error
            pass

    return {"error": f"Unknown operation: {op}"}


def _named_ranges_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    wb, err = _open_workbook_lib(resolved, display_path)
    if err:
        return {"error": err}

    op = kwargs.get("operation", "list")
    range_name = kwargs.get("range_name")

    if op == "list":
        names = []
        for dn in wb.defined_names.values():
            names.append({"name": dn.name, "refers_to": dn.attr_text})
        wb.close()
        return {"result": "Listed named ranges", "names": names}

    if op == "add":
        from openpyxl.workbook.defined_name import DefinedName

        cell_range = kwargs.get("cell_range")
        sheet_name = kwargs.get("sheet_name") or (wb.active.title if wb.active else "Sheet")
        if not range_name or not cell_range:
            wb.close()
            return {"error": "range_name and cell_range required for add"}
        dn = DefinedName(range_name, attr_text=f"'{sheet_name}'!{cell_range}")
        wb.defined_names.add(dn)
        wb.save(resolved)
        wb.close()
        return {"result": f"Added named range '{range_name}'", "path": display_path}

    if op == "delete":
        if not range_name:
            wb.close()
            return {"error": "range_name required for delete"}
        if range_name in wb.defined_names:
            del wb.defined_names[range_name]
        wb.save(resolved)
        wb.close()
        return {"result": f"Deleted named range '{range_name}'", "path": display_path}

    wb.close()
    return {"error": f"Unknown operation: {op}. Use 'list', 'add', or 'delete'"}


# --- data_validation ---


def _data_validation_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        ws, err = _get_sheet_com(wb, kwargs.get("sheet_name"))
        if err:
            return {"error": err}

        cell_range = kwargs.get("cell_range")
        validation = kwargs.get("validation") or {}
        if not cell_range or not validation:
            return {"error": "cell_range and validation required"}

        rng = ws.Range(cell_range)
        v_type = validation.get("type", "list")
        type_map = {"list": 3, "whole": 1, "decimal": 2, "date": 4, "textLength": 6}
        xl_type = type_map.get(v_type, 3)

        formula1 = validation.get("formula1", "")
        rng.Validation.Delete()
        rng.Validation.Add(Type=xl_type, Formula1=formula1)

        if validation.get("error_message"):
            rng.Validation.ErrorMessage = validation["error_message"]

        wb.Save()
    finally:
        wb.Close(SaveChanges=False)

    return {"result": f"Added data validation to {cell_range} in {display_path}", "path": display_path}


def _data_validation_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    from openpyxl.worksheet.datavalidation import DataValidation

    wb, err = _open_workbook_lib(resolved, display_path)
    if err:
        return {"error": err}

    ws, err = _get_sheet_lib(wb, kwargs.get("sheet_name"))
    if err:
        wb.close()
        return {"error": err}

    cell_range = kwargs.get("cell_range")
    validation = kwargs.get("validation") or {}
    if not cell_range or not validation:
        wb.close()
        return {"error": "cell_range and validation required"}

    dv = DataValidation(
        type=validation.get("type", "list"),
        formula1=validation.get("formula1", ""),
        allow_blank=validation.get("allow_blank", True),
    )
    if validation.get("error_message"):
        dv.error = validation["error_message"]
    dv.sqref = cell_range
    ws.add_data_validation(dv)
    wb.save(resolved)
    wb.close()
    return {"result": f"Added data validation to {cell_range} in {display_path}", "path": display_path}


# --- conditional_format ---


def _conditional_format_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        ws, err = _get_sheet_com(wb, kwargs.get("sheet_name"))
        if err:
            return {"error": err}

        cell_range = kwargs.get("cell_range")
        rule = kwargs.get("rule") or {}
        if not cell_range or not rule:
            return {"error": "cell_range and rule required"}

        rng = ws.Range(cell_range)
        operator = rule.get("operator", "greaterThan")
        formula = rule.get("formula", "0")

        op_map = {
            "greaterThan": 5,
            "lessThan": 6,
            "equal": 3,
            "notEqual": 4,
            "between": 1,
            "greaterThanOrEqual": 7,
            "lessThanOrEqual": 8,
        }
        xl_op = op_map.get(operator, 5)

        rng.FormatConditions.Add(Type=1, Operator=xl_op, Formula1=formula)
        fc = rng.FormatConditions(rng.FormatConditions.Count)

        fmt = rule.get("format") or {}
        if fmt.get("font_color"):
            fc.Font.Color = _parse_color_int(fmt["font_color"])
        if fmt.get("fill_color"):
            fc.Interior.Color = _parse_color_int(fmt["fill_color"])
        if fmt.get("bold") is not None:
            fc.Font.Bold = fmt["bold"]

        wb.Save()
    finally:
        wb.Close(SaveChanges=False)

    return {"result": f"Added conditional format to {cell_range} in {display_path}", "path": display_path}


def _conditional_format_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Font, PatternFill

    wb, err = _open_workbook_lib(resolved, display_path)
    if err:
        return {"error": err}

    ws, err = _get_sheet_lib(wb, kwargs.get("sheet_name"))
    if err:
        wb.close()
        return {"error": err}

    cell_range = kwargs.get("cell_range")
    rule = kwargs.get("rule") or {}
    if not cell_range or not rule:
        wb.close()
        return {"error": "cell_range and rule required"}

    operator = rule.get("operator", "greaterThan")
    formula = rule.get("formula", "0")
    fmt = rule.get("format") or {}

    if operator not in _VALID_CF_OPERATORS:
        wb.close()
        supported = ", ".join(_VALID_CF_OPERATORS)
        return {"error": f"Unknown conditional format operator: '{operator}'. Supported: {supported}"}

    try:
        font = Font(color=_normalize_hex_color(fmt["font_color"])) if fmt.get("font_color") else None
        fill = (
            PatternFill(start_color=_normalize_hex_color(fmt["fill_color"]), fill_type="solid")
            if fmt.get("fill_color")
            else None
        )

        cf_rule = CellIsRule(operator=operator, formula=[formula], font=font, fill=fill)
        ws.conditional_formatting.add(cell_range, cf_rule)
        wb.save(resolved)
    finally:
        wb.close()
    return {"result": f"Added conditional format to {cell_range} in {display_path}", "path": display_path}


# --- comments ---


def _comments_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        ws, err = _get_sheet_com(wb, kwargs.get("sheet_name"))
        if err:
            return {"error": err}

        op = kwargs.get("operation", "read")
        cell_range = kwargs.get("cell_range")

        if op == "add":
            if not cell_range or not kwargs.get("comment_text"):
                return {"error": "cell_range and comment_text required for add"}
            cell = ws.Range(cell_range)
            if cell.Comment is not None:
                cell.Comment.Delete()
            cell.AddComment(kwargs["comment_text"])
            wb.Save()
            return {"result": f"Added comment to {cell_range}", "path": display_path}

        if op == "read":
            comments: list[dict[str, str]] = []
            used = ws.UsedRange
            if used:
                for r in range(1, used.Rows.Count + 1):
                    for c in range(1, used.Columns.Count + 1):
                        cell = used.Cells(r, c)
                        if cell.Comment is not None:
                            comments.append(
                                {
                                    "cell": cell.Address.replace("$", ""),
                                    "text": cell.Comment.Text(),
                                    "author": cell.Comment.Author or "",
                                }
                            )
            return {"result": "Read comments", "comments": comments}

        if op == "delete":
            if not cell_range:
                return {"error": "cell_range required for delete"}
            cell = ws.Range(cell_range)
            if cell.Comment is not None:
                cell.Comment.Delete()
            wb.Save()
            return {"result": f"Deleted comment from {cell_range}", "path": display_path}
    finally:
        try:
            wb.Close(SaveChanges=False)
        except Exception:  # COM cleanup; must not mask the primary error
            pass

    return {"error": f"Unknown operation: {op}. Use 'add', 'read', or 'delete'"}


def _comments_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    from openpyxl.comments import Comment

    wb, err = _open_workbook_lib(resolved, display_path)
    if err:
        return {"error": err}

    ws, err = _get_sheet_lib(wb, kwargs.get("sheet_name"))
    if err:
        wb.close()
        return {"error": err}

    op = kwargs.get("operation", "read")
    cell_range = kwargs.get("cell_range")

    if op == "add":
        if not cell_range or not kwargs.get("comment_text"):
            wb.close()
            return {"error": "cell_range and comment_text required for add"}
        author = kwargs.get("author", "Anteroom")
        ws[cell_range].comment = Comment(kwargs["comment_text"], author)
        wb.save(resolved)
        wb.close()
        return {"result": f"Added comment to {cell_range}", "path": display_path}

    if op == "read":
        comments: list[dict[str, str]] = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment:
                    comments.append(
                        {
                            "cell": cell.coordinate,
                            "text": cell.comment.text,
                            "author": cell.comment.author or "",
                        }
                    )
        wb.close()
        return {"result": "Read comments", "comments": comments}

    if op == "delete":
        if not cell_range:
            wb.close()
            return {"error": "cell_range required for delete"}
        ws[cell_range].comment = None
        wb.save(resolved)
        wb.close()
        return {"result": f"Deleted comment from {cell_range}", "path": display_path}

    wb.close()
    return {"error": f"Unknown operation: {op}. Use 'add', 'read', or 'delete'"}


# --- hyperlinks ---


def _hyperlinks_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        ws, err = _get_sheet_com(wb, kwargs.get("sheet_name"))
        if err:
            return {"error": err}

        op = kwargs.get("operation", "add")
        cell_range = kwargs.get("cell_range")

        if op == "add":
            url = kwargs.get("url")
            if not cell_range or not url:
                return {"error": "cell_range and url required for add"}
            url_err = _validate_url(url)
            if url_err:
                return {"error": url_err}
            display_text = kwargs.get("display_text", url)
            ws.Hyperlinks.Add(Anchor=ws.Range(cell_range), Address=url, TextToDisplay=display_text)
            wb.Save()
            return {"result": f"Added hyperlink to {cell_range}", "path": display_path}

        if op == "read":
            links: list[dict[str, str]] = []
            for i in range(1, ws.Hyperlinks.Count + 1):
                hl = ws.Hyperlinks(i)
                links.append(
                    {
                        "cell": hl.Range.Address.replace("$", ""),
                        "url": hl.Address or "",
                        "display": hl.TextToDisplay or "",
                    }
                )
            return {"result": "Read hyperlinks", "hyperlinks": links}
    finally:
        try:
            wb.Close(SaveChanges=False)
        except Exception:  # COM cleanup; must not mask the primary error
            pass

    return {"error": f"Unknown operation: {op}. Use 'add' or 'read'"}


def _hyperlinks_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    wb, err = _open_workbook_lib(resolved, display_path)
    if err:
        return {"error": err}

    ws, err = _get_sheet_lib(wb, kwargs.get("sheet_name"))
    if err:
        wb.close()
        return {"error": err}

    op = kwargs.get("operation", "add")
    cell_range = kwargs.get("cell_range")

    if op == "add":
        url = kwargs.get("url")
        if not cell_range or not url:
            wb.close()
            return {"error": "cell_range and url required for add"}
        url_err = _validate_url(url)
        if url_err:
            wb.close()
            return {"error": url_err}
        display_text = kwargs.get("display_text", url)
        ws[cell_range].hyperlink = url
        ws[cell_range].value = display_text
        wb.save(resolved)
        wb.close()
        return {"result": f"Added hyperlink to {cell_range}", "path": display_path}

    if op == "read":
        links: list[dict[str, str]] = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.hyperlink:
                    links.append(
                        {
                            "cell": cell.coordinate,
                            "url": cell.hyperlink.target or "",
                            "display": str(cell.value or ""),
                        }
                    )
        wb.close()
        return {"result": "Read hyperlinks", "hyperlinks": links}

    wb.close()
    return {"error": f"Unknown operation: {op}. Use 'add' or 'read'"}


# --- images ---


def _images_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        ws, err = _get_sheet_com(wb, kwargs.get("sheet_name"))
        if err:
            return {"error": err}

        image_path = kwargs.get("image_path")
        if not image_path:
            return {"error": "image_path is required"}

        wd = kwargs.get("working_dir", os.getcwd())
        img_resolved, img_err = validate_path(image_path, wd)
        if img_err:
            return {"error": img_err}
        if not os.path.isfile(img_resolved):
            return {"error": f"Image not found: {image_path}"}

        anchor_cell = kwargs.get("anchor_cell", "A1")
        rng = ws.Range(anchor_cell)
        ws.Shapes.AddPicture(
            os.path.abspath(img_resolved),
            LinkToFile=False,
            SaveWithDocument=True,
            Left=rng.Left,
            Top=rng.Top,
            Width=-1,
            Height=-1,
        )

        wb.Save()
    finally:
        wb.Close(SaveChanges=False)

    return {"result": f"Inserted image at {anchor_cell} in {display_path}", "path": display_path}


def _images_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    from openpyxl.drawing.image import Image

    wb, err = _open_workbook_lib(resolved, display_path)
    if err:
        return {"error": err}

    ws, err = _get_sheet_lib(wb, kwargs.get("sheet_name"))
    if err:
        wb.close()
        return {"error": err}

    image_path = kwargs.get("image_path")
    if not image_path:
        wb.close()
        return {"error": "image_path is required"}

    wd = kwargs.get("working_dir", os.getcwd())
    img_resolved, img_err = validate_path(image_path, wd)
    if img_err:
        wb.close()
        return {"error": img_err}
    if not os.path.isfile(img_resolved):
        wb.close()
        return {"error": f"Image not found: {image_path}"}

    anchor_cell = kwargs.get("anchor_cell", "A1")
    img = Image(img_resolved)
    ws.add_image(img, anchor_cell)
    wb.save(resolved)
    wb.close()
    return {"result": f"Inserted image at {anchor_cell} in {display_path}", "path": display_path}


# --- protect ---


def _protect_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        op = kwargs.get("operation", "enable")

        if op == "disable":
            password = kwargs.get("password", "")
            sheet_name = kwargs.get("sheet_name")
            if sheet_name:
                ws, err = _get_sheet_com(wb, sheet_name)
                if err:
                    return {"error": err}
                ws.Unprotect(Password=password)
            else:
                wb.Unprotect(Password=password)
            wb.Save()
            return {"result": f"Removed protection from {display_path}", "path": display_path}

        # enable
        password = kwargs.get("password", "")
        sheet_name = kwargs.get("sheet_name")
        if sheet_name:
            ws, err = _get_sheet_com(wb, sheet_name)
            if err:
                return {"error": err}
            ws.Protect(Password=password)
        else:
            wb.Protect(Password=password)
        wb.Save()
    finally:
        try:
            wb.Close(SaveChanges=False)
        except Exception:  # COM cleanup; must not mask the primary error
            pass

    return {"result": f"Protected {display_path}", "path": display_path}


def _protect_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    wb, err = _open_workbook_lib(resolved, display_path)
    if err:
        return {"error": err}

    op = kwargs.get("operation", "enable")
    password = kwargs.get("password")
    sheet_name = kwargs.get("sheet_name")

    if op == "disable":
        if sheet_name:
            ws, err = _get_sheet_lib(wb, sheet_name)
            if err:
                wb.close()
                return {"error": err}
            ws.protection.sheet = False
            ws.protection.password = ""
        else:
            wb.security.workbookPassword = ""
            wb.security.lockStructure = False
        wb.save(resolved)
        wb.close()
        return {"result": f"Removed protection from {display_path}", "path": display_path}

    # enable
    if sheet_name:
        ws, err = _get_sheet_lib(wb, sheet_name)
        if err:
            wb.close()
            return {"error": err}
        ws.protection.sheet = True
        if password:
            ws.protection.password = password
    else:
        wb.security.lockStructure = True
        if password:
            wb.security.workbookPassword = password

    wb.save(resolved)
    wb.close()
    return {"result": f"Protected {display_path}", "path": display_path}


# --- group_rows_cols ---


def _group_rows_cols_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        ws, err = _get_sheet_com(wb, kwargs.get("sheet_name"))
        if err:
            return {"error": err}

        start = kwargs.get("start")
        end = kwargs.get("end")
        axis = kwargs.get("axis", "rows")
        op = kwargs.get("operation", "group")

        if start is None or end is None:
            return {"error": "start and end required"}

        if axis == "rows":
            rng = ws.Rows(f"{start}:{end}")
        else:
            rng = ws.Columns(f"{start}:{end}")

        if op == "ungroup":
            rng.Ungroup()
        else:
            rng.Group()

        wb.Save()
    finally:
        wb.Close(SaveChanges=False)

    return {"result": f"{'Grouped' if op != 'ungroup' else 'Ungrouped'} {axis} {start}-{end}", "path": display_path}


def _group_rows_cols_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    wb, err = _open_workbook_lib(resolved, display_path)
    if err:
        return {"error": err}

    ws, err = _get_sheet_lib(wb, kwargs.get("sheet_name"))
    if err:
        wb.close()
        return {"error": err}

    start = kwargs.get("start")
    end = kwargs.get("end")
    axis = kwargs.get("axis", "rows")
    op = kwargs.get("operation", "group")

    if start is None or end is None:
        wb.close()
        return {"error": "start and end required"}

    if axis == "rows":
        if op == "ungroup":
            ws.row_dimensions.group(start, end, outline_level=0)
        else:
            ws.row_dimensions.group(start, end)
    else:
        from openpyxl.utils import get_column_letter

        start_letter = get_column_letter(start) if isinstance(start, int) else start
        end_letter = get_column_letter(end) if isinstance(end, int) else end
        if op == "ungroup":
            ws.column_dimensions.group(start_letter, end_letter, outline_level=0)
        else:
            ws.column_dimensions.group(start_letter, end_letter)

    wb.save(resolved)
    wb.close()
    return {"result": f"{'Grouped' if op != 'ungroup' else 'Ungrouped'} {axis} {start}-{end}", "path": display_path}


# --- print_settings ---


def _print_settings_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        ws, err = _get_sheet_com(wb, kwargs.get("sheet_name"))
        if err:
            return {"error": err}

        ps = kwargs.get("page_setup") or {}
        setup = ws.PageSetup

        if ps.get("orientation"):
            setup.Orientation = 2 if ps["orientation"] == "landscape" else 1
        if ps.get("paper_size"):
            setup.PaperSize = ps["paper_size"]
        if ps.get("fit_to_width") is not None:
            setup.FitToPagesWide = ps["fit_to_width"]
        if ps.get("fit_to_height") is not None:
            setup.FitToPagesTall = ps["fit_to_height"]
        if ps.get("header"):
            setup.CenterHeader = ps["header"]
        if ps.get("footer"):
            setup.CenterFooter = ps["footer"]

        margins = ps.get("margins") or {}
        if margins.get("top") is not None:
            setup.TopMargin = margins["top"]
        if margins.get("bottom") is not None:
            setup.BottomMargin = margins["bottom"]
        if margins.get("left") is not None:
            setup.LeftMargin = margins["left"]
        if margins.get("right") is not None:
            setup.RightMargin = margins["right"]

        wb.Save()
    finally:
        wb.Close(SaveChanges=False)

    return {"result": f"Updated print settings in {display_path}", "path": display_path}


def _print_settings_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    wb, err = _open_workbook_lib(resolved, display_path)
    if err:
        return {"error": err}

    ws, err = _get_sheet_lib(wb, kwargs.get("sheet_name"))
    if err:
        wb.close()
        return {"error": err}

    ps = kwargs.get("page_setup") or {}

    if ps.get("orientation"):
        ws.page_setup.orientation = ps["orientation"]
    if ps.get("paper_size"):
        ws.page_setup.paperSize = ps["paper_size"]
    if ps.get("fit_to_width") is not None:
        ws.page_setup.fitToWidth = ps["fit_to_width"]
    if ps.get("fit_to_height") is not None:
        ws.page_setup.fitToHeight = ps["fit_to_height"]
    if ps.get("header"):
        ws.oddHeader.center.text = ps["header"]
    if ps.get("footer"):
        ws.oddFooter.center.text = ps["footer"]

    margins = ps.get("margins") or {}
    if margins.get("top") is not None:
        ws.page_margins.top = margins["top"]
    if margins.get("bottom") is not None:
        ws.page_margins.bottom = margins["bottom"]
    if margins.get("left") is not None:
        ws.page_margins.left = margins["left"]
    if margins.get("right") is not None:
        ws.page_margins.right = margins["right"]

    wb.save(resolved)
    wb.close()
    return {"result": f"Updated print settings in {display_path}", "path": display_path}


# --- charts ---


def _charts_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        ws, err = _get_sheet_com(wb, kwargs.get("sheet_name"))
        if err:
            return {"error": err}

        data_range = kwargs.get("data_range")
        if not data_range:
            return {"error": "data_range is required for charts"}

        chart_type = kwargs.get("chart_type", "column")
        chart_title = kwargs.get("chart_title", "")
        anchor_cell = kwargs.get("anchor_cell", "E2")

        type_map = {
            "column": 51,
            "bar": 57,
            "line": 4,
            "pie": 5,
            "scatter": -4169,
            "area": 1,
        }
        if chart_type not in type_map:
            return {"error": f"Unknown chart_type: '{chart_type}'. Supported: {', '.join(_VALID_CHART_TYPES)}"}
        xl_type = type_map[chart_type]

        rng = ws.Range(data_range)
        anchor = ws.Range(anchor_cell)

        chart_obj = ws.Shapes.AddChart2(Style=-1, XlChartType=xl_type, Left=anchor.Left, Top=anchor.Top)
        chart_obj.Chart.SetSourceData(Source=rng)
        if chart_title:
            chart_obj.Chart.HasTitle = True
            chart_obj.Chart.ChartTitle.Text = chart_title

        wb.Save()
    finally:
        wb.Close(SaveChanges=False)

    return {"result": f"Created {chart_type} chart in {display_path}", "path": display_path}


def _charts_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, Reference, ScatterChart

    wb, err = _open_workbook_lib(resolved, display_path)
    if err:
        return {"error": err}

    ws, err = _get_sheet_lib(wb, kwargs.get("sheet_name"))
    if err:
        wb.close()
        return {"error": err}

    data_range = kwargs.get("data_range")
    if not data_range:
        wb.close()
        return {"error": "data_range is required for charts"}

    chart_type = kwargs.get("chart_type", "column")
    chart_title = kwargs.get("chart_title", "")
    anchor_cell = kwargs.get("anchor_cell", "E2")

    type_map: dict[str, Any] = {
        "column": BarChart,
        "bar": BarChart,
        "line": LineChart,
        "pie": PieChart,
        "scatter": ScatterChart,
        "area": AreaChart,
    }
    if chart_type not in type_map:
        wb.close()
        return {"error": f"Unknown chart_type: '{chart_type}'. Supported: {', '.join(_VALID_CHART_TYPES)}"}
    chart_cls = type_map[chart_type]

    try:
        chart = chart_cls()
        if chart_title:
            chart.title = chart_title

        # Parse range like "A1:D10" to Reference
        from openpyxl.utils.cell import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(data_range)
        data = Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
        chart.add_data(data, titles_from_data=True)

        ws.add_chart(chart, anchor_cell)
        wb.save(resolved)
    finally:
        wb.close()
    return {"result": f"Created {chart_type} chart in {display_path}", "path": display_path}


# --- COM-only actions ---


def _export_pdf_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        output_path = kwargs.get("output_path")
        if not output_path:
            output_path = os.path.splitext(resolved)[0] + ".pdf"
        else:
            wd = kwargs.get("working_dir", os.getcwd())
            out_resolved, out_err = validate_path(output_path, wd)
            if out_err:
                return {"error": out_err}
            output_path = out_resolved

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.ExportAsFixedFormat(Type=0, Filename=os.path.abspath(output_path))
    finally:
        wb.Close(SaveChanges=False)

    return {"result": f"Exported PDF to {output_path}", "path": display_path}


def _export_pdf_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    return {"error": "Action 'export_pdf' requires Windows with Office installed (COM backend)"}


def _sort_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        ws, err = _get_sheet_com(wb, kwargs.get("sheet_name"))
        if err:
            return {"error": err}

        cell_range = kwargs.get("cell_range")
        sort_column = kwargs.get("sort_column")
        if not cell_range or not sort_column:
            return {"error": "cell_range and sort_column required for sort"}

        ascending = kwargs.get("ascending", True)
        order = 1 if ascending else 2

        rng = ws.Range(cell_range)
        key = ws.Range(sort_column + "1")
        rng.Sort(Key1=key, Order1=order, Header=1)
        wb.Save()
    finally:
        wb.Close(SaveChanges=False)

    return {"result": f"Sorted {cell_range} by {sort_column} in {display_path}", "path": display_path}


def _sort_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    return {"error": "Action 'sort' requires Windows with Office installed (COM backend)"}


def _pivot_tables_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        ws, err = _get_sheet_com(wb, kwargs.get("sheet_name"))
        if err:
            return {"error": err}

        data_range = kwargs.get("data_range")
        if not data_range:
            return {"error": "data_range is required for pivot_tables"}

        # Create pivot table on a new sheet
        dest_ws = wb.Worksheets.Add(After=wb.Worksheets(wb.Worksheets.Count))
        dest_ws.Name = kwargs.get("range_name", "PivotTable")

        src_range = ws.Range(data_range)
        pivot_cache = wb.PivotCaches().Create(SourceType=1, SourceData=src_range)
        dest_cell = dest_ws.Range("A3")
        pivot_cache.CreatePivotTable(TableDestination=dest_cell, TableName="PivotTable1")

        wb.Save()
    finally:
        wb.Close(SaveChanges=False)

    return {"result": f"Created pivot table in {display_path}", "path": display_path}


def _pivot_tables_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    return {"error": "Action 'pivot_tables' requires Windows with Office installed (COM backend)"}


def _sparklines_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        ws, err = _get_sheet_com(wb, kwargs.get("sheet_name"))
        if err:
            return {"error": err}

        data_range = kwargs.get("data_range")
        anchor_cell = kwargs.get("anchor_cell")
        if not data_range or not anchor_cell:
            return {"error": "data_range and anchor_cell required for sparklines"}

        chart_type = kwargs.get("chart_type", "line")
        type_map = {"line": 1, "column": 2, "win_loss": 3}
        xl_type = type_map.get(chart_type, 1)

        ws.Range(anchor_cell).SparklineGroups.Add(Type=xl_type, SourceData=data_range)
        wb.Save()
    finally:
        wb.Close(SaveChanges=False)

    return {"result": f"Added sparkline at {anchor_cell} in {display_path}", "path": display_path}


def _sparklines_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    return {"error": "Action 'sparklines' requires Windows with Office installed (COM backend)"}


def _slicers_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        ws, err = _get_sheet_com(wb, kwargs.get("sheet_name"))
        if err:
            return {"error": err}

        # Slicers require an existing pivot table or table
        range_name = kwargs.get("range_name")
        if not range_name:
            return {"error": "range_name (pivot table or table name) required for slicers"}

        anchor_cell = kwargs.get("anchor_cell", "H1")
        anchor = ws.Range(anchor_cell)

        slicer_cache = wb.SlicerCaches.Add2(
            wb.PivotTables(range_name) if hasattr(wb, "PivotTables") else ws.ListObjects(range_name),
            range_name,
        )
        slicer_cache.Slicers.Add(ws, Left=anchor.Left, Top=anchor.Top)
        wb.Save()
    finally:
        wb.Close(SaveChanges=False)

    return {"result": f"Added slicer for {range_name} in {display_path}", "path": display_path}


def _slicers_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    return {"error": "Action 'slicers' requires Windows with Office installed (COM backend)"}


# ---------------------------------------------------------------------------
# Library backend (openpyxl) — create/read/edit
# ---------------------------------------------------------------------------


def _create_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    import openpyxl as _openpyxl

    sheets: list[dict[str, Any]] = kwargs.get("sheets") or []
    if not sheets:
        return {"error": "sheets is required for create action"}
    if len(sheets) > _MAX_CONTENT_BLOCKS:
        return {"error": f"Too many sheets (max {_MAX_CONTENT_BLOCKS})"}

    wb = _openpyxl.Workbook()
    try:
        default_sheet = wb.active
        if default_sheet is not None:
            wb.remove(default_sheet)

        total_rows = 0
        for sheet_def in sheets:
            name = sheet_def.get("name", "Sheet")
            ws = wb.create_sheet(title=str(name))
            headers: list[str] = sheet_def.get("headers") or []
            rows: list[list[Any]] = sheet_def.get("rows") or []

            if headers:
                ws.append(headers)
                total_rows += 1

            for row in rows:
                if total_rows >= _MAX_ROWS:
                    break
                ws.append(row)
                total_rows += 1

        if total_rows >= _MAX_ROWS:
            return {"error": f"Too many rows (max {_MAX_ROWS})"}

        os.makedirs(os.path.dirname(resolved) or ".", exist_ok=True)
        wb.save(resolved)
    finally:
        wb.close()
    return {
        "result": f"Created {display_path}",
        "path": display_path,
        "sheets_created": len(sheets),
        "total_rows": total_rows,
    }


def _read_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    import openpyxl as _openpyxl

    sheet_name: str | None = kwargs.get("sheet_name")
    cell_range: str | None = kwargs.get("cell_range")

    # Open twice: data_only=True for computed values, data_only=False for formulas
    wb_val, err = _open_workbook_lib(resolved, display_path, read_only=True)
    if err:
        return {"error": err}

    try:
        wb_fmt = _openpyxl.load_workbook(resolved, read_only=False, data_only=False)
    except Exception:
        wb_fmt = None

    ws_val, err = _get_sheet_lib(wb_val, sheet_name)
    if err:
        wb_val.close()
        if wb_fmt:
            wb_fmt.close()
        return {"error": err}

    ws_fmt = None
    if wb_fmt:
        ws_fmt, _ = _get_sheet_lib(wb_fmt, sheet_name)

    output_rows: list[list[Any]] = []
    formulas: dict[str, str] = {}
    formatting: dict[str, dict[str, Any]] = {}

    try:
        if cell_range:
            for row in ws_val[cell_range]:
                output_rows.append([cell.value for cell in row])
                if len(output_rows) >= _MAX_ROWS:
                    break
        else:
            for row in ws_val.iter_rows():
                output_rows.append([cell.value for cell in row])
                if len(output_rows) >= _MAX_ROWS:
                    break
    except Exception as exc:
        wb_val.close()
        if wb_fmt:
            wb_fmt.close()
        return {"error": f"Unable to read range from: {display_path} ({type(exc).__name__}: {exc})"}

    # Collect formulas from the non-data_only workbook
    if ws_fmt:
        try:
            for row in ws_fmt.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formulas[cell.coordinate] = cell.value
        except Exception:
            pass

    # Collect cell formatting for non-default cells
    if ws_fmt:
        try:
            from openpyxl.styles import DEFAULT_FONT

            for row in ws_fmt.iter_rows():
                for cell in row:
                    if cell.value is None and not isinstance(cell.value, (int, float)):
                        if cell.value is None:
                            continue
                    fmt_info: dict[str, Any] = {}
                    f = cell.font
                    if f and f != DEFAULT_FONT:
                        if f.bold:
                            fmt_info["bold"] = True
                        if f.italic:
                            fmt_info["italic"] = True
                        if f.underline:
                            fmt_info["underline"] = True
                        if f.name and f.name != DEFAULT_FONT.name:
                            fmt_info["font"] = f.name
                        if f.size and f.size != DEFAULT_FONT.size:
                            fmt_info["size"] = f.size
                        if f.color and f.color.rgb and f.color.rgb != "00000000":
                            color_val = str(f.color.rgb)
                            if color_val not in ("00000000", "FF000000"):
                                fmt_info["font_color"] = f"#{color_val[-6:]}"
                    fill = cell.fill
                    if fill and fill.fgColor and fill.fgColor.rgb:
                        fill_rgb = str(fill.fgColor.rgb)
                        if fill_rgb not in ("00000000", "FFFFFFFF", "00000000"):
                            fmt_info["fill"] = f"#{fill_rgb[-6:]}"
                    if cell.number_format and cell.number_format != "General":
                        fmt_info["number_format"] = cell.number_format
                    al = cell.alignment
                    if al:
                        if al.horizontal and al.horizontal != "general":
                            fmt_info["align"] = al.horizontal
                        if al.wrap_text:
                            fmt_info["wrap"] = True
                    if fmt_info:
                        formatting[cell.coordinate] = fmt_info
        except Exception:
            pass

    # Collect merged ranges
    merged_ranges: list[str] = []
    if ws_fmt:
        try:
            for mr in ws_fmt.merged_cells.ranges:
                merged_ranges.append(str(mr))
        except Exception:
            pass

    # Collect column widths and row heights
    col_widths: dict[str, float] = {}
    row_heights: dict[int, float] = {}
    if ws_fmt:
        try:
            for col_letter, dim in ws_fmt.column_dimensions.items():
                if dim.width is not None and dim.width != 8.0:
                    col_widths[str(col_letter)] = round(dim.width, 2)
            for row_num, dim in ws_fmt.row_dimensions.items():
                if dim.height is not None and dim.height != 15.0:
                    row_heights[int(row_num)] = round(dim.height, 2)
        except Exception:
            pass

    # Collect data validations
    validations: list[dict[str, Any]] = []
    if ws_fmt:
        try:
            for dv in ws_fmt.data_validations.dataValidation:
                v_info: dict[str, Any] = {"range": str(dv.sqref)}
                if dv.type:
                    v_info["type"] = dv.type
                if dv.formula1:
                    v_info["formula1"] = str(dv.formula1)
                if dv.formula2:
                    v_info["formula2"] = str(dv.formula2)
                if dv.operator:
                    v_info["operator"] = dv.operator
                validations.append(v_info)
        except Exception:
            pass

    # Collect conditional formatting rules
    cf_rules: list[dict[str, Any]] = []
    if ws_fmt:
        try:
            for cf_range, rules in ws_fmt.conditional_formatting:
                for rule in rules:
                    cf_info: dict[str, Any] = {"range": str(cf_range)}
                    if hasattr(rule, "type"):
                        cf_info["type"] = rule.type
                    if hasattr(rule, "operator") and rule.operator:
                        cf_info["operator"] = rule.operator
                    if hasattr(rule, "formula") and rule.formula:
                        cf_info["formula"] = [str(f) for f in rule.formula]
                    cf_rules.append(cf_info)
        except Exception:
            pass

    # Collect named ranges
    named_ranges: list[dict[str, str]] = []
    if wb_fmt:
        try:
            for dn in wb_fmt.defined_names.values():
                named_ranges.append({"name": dn.name, "refers_to": dn.attr_text})
        except Exception:
            pass

    # Collect freeze pane position
    freeze_pane: str | None = None
    if ws_fmt:
        try:
            if ws_fmt.freeze_panes:
                freeze_pane = str(ws_fmt.freeze_panes)
        except Exception:
            pass

    # Collect auto-filter state
    auto_filter_ref: str | None = None
    if ws_fmt:
        try:
            if ws_fmt.auto_filter and ws_fmt.auto_filter.ref:
                auto_filter_ref = str(ws_fmt.auto_filter.ref)
        except Exception:
            pass

    # Collect hidden sheets
    hidden_sheets: list[str] = []
    if wb_fmt:
        try:
            for sn in wb_fmt.sheetnames:
                s = wb_fmt[sn]
                if s.sheet_state and s.sheet_state != "visible":
                    hidden_sheets.append(f"{sn} ({s.sheet_state})")
        except Exception:
            pass

    sheet_title = ws_val.title
    sheets_available = list(wb_val.sheetnames)
    wb_val.close()
    if wb_fmt:
        wb_fmt.close()

    content = json.dumps(output_rows, ensure_ascii=False, default=str)
    if len(content) > _MAX_OUTPUT:
        content = content[:_MAX_OUTPUT] + "\n... (truncated)"

    result: dict[str, Any] = {
        "content": content,
        "sheet": sheet_title,
        "sheets_available": sheets_available,
        "rows_read": len(output_rows),
    }

    if formulas:
        result["formulas"] = formulas
    if formatting:
        result["formatting"] = formatting
    if merged_ranges:
        result["merged_ranges"] = merged_ranges
    if col_widths:
        result["column_widths"] = col_widths
    if row_heights:
        result["row_heights"] = row_heights
    if validations:
        result["data_validations"] = validations
    if cf_rules:
        result["conditional_formatting"] = cf_rules
    if named_ranges:
        result["named_ranges"] = named_ranges
    if freeze_pane:
        result["freeze_pane"] = freeze_pane
    if auto_filter_ref:
        result["auto_filter"] = auto_filter_ref
    if hidden_sheets:
        result["hidden_sheets"] = hidden_sheets

    return result


def _edit_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    wb, err = _open_workbook_lib(resolved, display_path)
    if err:
        return {"error": err}

    sheet_name: str | None = kwargs.get("sheet_name")
    updates: list[dict[str, Any]] = kwargs.get("updates") or []
    append_rows: list[list[Any]] = kwargs.get("append_rows") or []
    add_sheets: list[dict[str, Any]] = kwargs.get("add_sheets") or []

    if not updates and not append_rows and not add_sheets:
        wb.close()
        return {"error": "Provide 'updates', 'append_rows', and/or 'add_sheets' for edit action"}

    if append_rows and len(append_rows) > _MAX_ROWS:
        wb.close()
        return {"error": f"Too many rows to append (max {_MAX_ROWS})"}

    cells_updated = 0
    rows_appended = 0
    sheets_added = 0

    if updates or append_rows:
        ws, err = _get_sheet_lib(wb, sheet_name)
        if err:
            wb.close()
            return {"error": err}

        for upd in updates:
            cell = upd.get("cell", "")
            value = upd.get("value")
            if not cell:
                continue
            ws[cell] = value
            cells_updated += 1

        for row in append_rows:
            ws.append(row)
            rows_appended += 1

    for sheet_def in add_sheets:
        name = sheet_def.get("name", "Sheet")
        ws_new = wb.create_sheet(title=str(name))
        rows = sheet_def.get("rows") or []
        for row in rows[:_MAX_ROWS]:
            ws_new.append(row)
        sheets_added += 1

    wb.save(resolved)
    wb.close()
    return {
        "result": f"Edited {display_path}",
        "path": display_path,
        "cells_updated": cells_updated,
        "rows_appended": rows_appended,
        "sheets_added": sheets_added,
    }


# ---------------------------------------------------------------------------
# template_fill — {{key}} replacement across all cells
# ---------------------------------------------------------------------------


def _template_fill_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    template_data: dict[str, Any] = kwargs.get("template_data") or {}
    if not template_data:
        return {"error": "template_data is required for template_fill action"}
    if len(template_data) > _MAX_EDIT_OPS:
        return {"error": f"Too many template keys (max {_MAX_EDIT_OPS})"}

    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        tokens_replaced = 0
        for key, value in template_data.items():
            token = "{{" + str(key) + "}}"
            val_str = str(value)
            for si in range(1, wb.Worksheets.Count + 1):
                ws = wb.Worksheets(si)
                used = ws.UsedRange
                if used is None:
                    continue
                cells = used.Find(What=token, LookIn=-4163, LookAt=2)  # xlValues, xlPart
                if cells is None:
                    continue
                first_addr = cells.Address
                while True:
                    cells.Value = _sanitize_cell_value(str(cells.Value).replace(token, val_str))
                    tokens_replaced += 1
                    cells = used.FindNext(cells)
                    if cells is None or cells.Address == first_addr:
                        break
        wb.Save()
    finally:
        wb.Close(SaveChanges=False)

    return {
        "result": f"Template fill completed on {display_path}",
        "path": display_path,
        "tokens_replaced": tokens_replaced,
        "keys_processed": len(template_data),
    }


def _template_fill_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    template_data: dict[str, Any] = kwargs.get("template_data") or {}
    if not template_data:
        return {"error": "template_data is required for template_fill action"}
    if len(template_data) > _MAX_EDIT_OPS:
        return {"error": f"Too many template keys (max {_MAX_EDIT_OPS})"}

    wb, err = _open_workbook_lib(resolved, display_path)
    if err:
        return {"error": err}

    tokens_replaced = 0
    for key, value in template_data.items():
        token = "{{" + str(key) + "}}"
        val_str = str(value)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and token in cell.value:
                        cell.value = cell.value.replace(token, val_str)
                        tokens_replaced += 1

    wb.save(resolved)
    wb.close()
    return {
        "result": f"Template fill completed on {display_path}",
        "path": display_path,
        "tokens_replaced": tokens_replaced,
        "keys_processed": len(template_data),
    }


# ---------------------------------------------------------------------------
# manage_sheets — rename, delete, copy, reorder, hide/unhide
# ---------------------------------------------------------------------------

_VALID_SHEET_OPS = ("rename", "delete", "copy", "reorder", "hide", "unhide")


def _manage_sheets_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    sheet_operations: list[dict[str, Any]] = kwargs.get("sheet_operations") or []
    if not sheet_operations:
        return {"error": "sheet_operations is required for manage_sheets action"}
    if len(sheet_operations) > _MAX_EDIT_OPS:
        return {"error": f"Too many sheet operations (max {_MAX_EDIT_OPS})"}

    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        ops_done = 0
        for op_def in sheet_operations:
            op = op_def.get("op", "")
            sheet = op_def.get("sheet", "")
            if op not in _VALID_SHEET_OPS:
                return {"error": f"Unknown sheet operation: '{op}'. Valid: {', '.join(_VALID_SHEET_OPS)}"}
            if not sheet:
                return {"error": "Each sheet operation requires a 'sheet' name"}

            ws, ws_err = _get_sheet_com(wb, sheet)
            if ws_err and op != "copy":
                return {"error": ws_err}

            if op == "rename":
                new_name = op_def.get("new_name")
                if not new_name:
                    return {"error": "rename requires 'new_name'"}
                ws.Name = str(new_name)
            elif op == "delete":
                if wb.Worksheets.Count <= 1:
                    return {"error": "Cannot delete the only sheet in a workbook"}
                ws.Delete()
            elif op == "copy":
                target_name = op_def.get("new_name", f"{sheet} Copy")
                ws.Copy(After=wb.Worksheets(wb.Worksheets.Count))
                wb.Worksheets(wb.Worksheets.Count).Name = str(target_name)
            elif op == "reorder":
                position = op_def.get("position", 1)
                position = max(1, min(position, wb.Worksheets.Count))
                if position == 1:
                    ws.Move(Before=wb.Worksheets(1))
                else:
                    ws.Move(After=wb.Worksheets(position - 1))
            elif op == "hide":
                ws.Visible = 0  # xlSheetHidden
            elif op == "unhide":
                ws.Visible = -1  # xlSheetVisible
            ops_done += 1

        wb.Save()
    finally:
        wb.Close(SaveChanges=False)

    return {
        "result": f"Completed {ops_done} sheet operation(s) on {display_path}",
        "path": display_path,
        "operations_completed": ops_done,
    }


def _manage_sheets_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    sheet_operations: list[dict[str, Any]] = kwargs.get("sheet_operations") or []
    if not sheet_operations:
        return {"error": "sheet_operations is required for manage_sheets action"}
    if len(sheet_operations) > _MAX_EDIT_OPS:
        return {"error": f"Too many sheet operations (max {_MAX_EDIT_OPS})"}

    wb, err = _open_workbook_lib(resolved, display_path)
    if err:
        return {"error": err}

    ops_done = 0
    for op_def in sheet_operations:
        op = op_def.get("op", "")
        sheet = op_def.get("sheet", "")
        if op not in _VALID_SHEET_OPS:
            wb.close()
            return {"error": f"Unknown sheet operation: '{op}'. Valid: {', '.join(_VALID_SHEET_OPS)}"}
        if not sheet:
            wb.close()
            return {"error": "Each sheet operation requires a 'sheet' name"}

        if sheet not in wb.sheetnames:
            wb.close()
            return {"error": f"Sheet '{sheet}' not found. Available: {list(wb.sheetnames)}"}

        ws = wb[sheet]

        if op == "rename":
            new_name = op_def.get("new_name")
            if not new_name:
                wb.close()
                return {"error": "rename requires 'new_name'"}
            ws.title = str(new_name)
        elif op == "delete":
            if len(wb.sheetnames) <= 1:
                wb.close()
                return {"error": "Cannot delete the only sheet in a workbook"}
            wb.remove(ws)
        elif op == "copy":
            target_name = op_def.get("new_name", f"{sheet} Copy")
            new_ws = wb.copy_worksheet(ws)
            new_ws.title = str(target_name)
        elif op == "reorder":
            position = op_def.get("position", 1)
            position = max(1, min(position, len(wb.sheetnames))) - 1
            wb.move_sheet(ws, offset=position - wb.sheetnames.index(sheet))
        elif op == "hide":
            ws.sheet_state = "hidden"
        elif op == "unhide":
            ws.sheet_state = "visible"
        ops_done += 1

    wb.save(resolved)
    wb.close()
    return {
        "result": f"Completed {ops_done} sheet operation(s) on {display_path}",
        "path": display_path,
        "operations_completed": ops_done,
    }


# ---------------------------------------------------------------------------
# resize — column width / row height control
# ---------------------------------------------------------------------------


def _resize_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    resize_ops: list[dict[str, Any]] = kwargs.get("resize_ops") or []
    if not resize_ops:
        return {"error": "resize_ops is required for resize action"}
    if len(resize_ops) > _MAX_EDIT_OPS:
        return {"error": f"Too many resize operations (max {_MAX_EDIT_OPS})"}

    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        ws, ws_err = _get_sheet_com(wb, kwargs.get("sheet_name"))
        if ws_err:
            return {"error": ws_err}

        resized = 0
        for op in resize_ops:
            target = op.get("target", "")
            index = op.get("index")
            size = op.get("size")
            if target not in ("column", "row"):
                return {"error": f"Invalid resize target: '{target}'. Use 'column' or 'row'"}
            if index is None or size is None:
                return {"error": "Each resize op requires 'index' and 'size'"}
            size = float(size)
            if size < 0:
                return {"error": "Size must be non-negative"}

            if target == "column":
                ws.Columns(str(index)).ColumnWidth = size
            else:
                ws.Rows(int(index)).RowHeight = size
            resized += 1

        wb.Save()
    finally:
        wb.Close(SaveChanges=False)

    return {
        "result": f"Resized {resized} column(s)/row(s) in {display_path}",
        "path": display_path,
        "resized": resized,
    }


def _resize_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    resize_ops: list[dict[str, Any]] = kwargs.get("resize_ops") or []
    if not resize_ops:
        return {"error": "resize_ops is required for resize action"}
    if len(resize_ops) > _MAX_EDIT_OPS:
        return {"error": f"Too many resize operations (max {_MAX_EDIT_OPS})"}

    wb, err = _open_workbook_lib(resolved, display_path)
    if err:
        return {"error": err}

    ws, ws_err = _get_sheet_lib(wb, kwargs.get("sheet_name"))
    if ws_err:
        wb.close()
        return {"error": ws_err}

    resized = 0
    for op in resize_ops:
        target = op.get("target", "")
        index = op.get("index")
        size = op.get("size")
        if target not in ("column", "row"):
            wb.close()
            return {"error": f"Invalid resize target: '{target}'. Use 'column' or 'row'"}
        if index is None or size is None:
            wb.close()
            return {"error": "Each resize op requires 'index' and 'size'"}
        size = float(size)
        if size < 0:
            wb.close()
            return {"error": "Size must be non-negative"}

        if target == "column":
            ws.column_dimensions[str(index).upper()].width = size
        else:
            ws.row_dimensions[int(index)].height = size
        resized += 1

    wb.save(resolved)
    wb.close()
    return {
        "result": f"Resized {resized} column(s)/row(s) in {display_path}",
        "path": display_path,
        "resized": resized,
    }


# ---------------------------------------------------------------------------
# insert_delete — bulk row/column insert/delete
# ---------------------------------------------------------------------------

_VALID_INSERT_DELETE_OPS = ("insert_rows", "delete_rows", "insert_cols", "delete_cols")


def _insert_delete_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    insert_delete_ops: list[dict[str, Any]] = kwargs.get("insert_delete_ops") or []
    if not insert_delete_ops:
        return {"error": "insert_delete_ops is required for insert_delete action"}
    if len(insert_delete_ops) > _MAX_EDIT_OPS:
        return {"error": f"Too many insert/delete operations (max {_MAX_EDIT_OPS})"}

    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        ws, ws_err = _get_sheet_com(wb, kwargs.get("sheet_name"))
        if ws_err:
            return {"error": ws_err}

        ops_done = 0
        for op_def in insert_delete_ops:
            op = op_def.get("op", "")
            index = op_def.get("index")
            count = op_def.get("count", 1)
            if op not in _VALID_INSERT_DELETE_OPS:
                return {"error": f"Unknown op: '{op}'. Valid: {', '.join(_VALID_INSERT_DELETE_OPS)}"}
            if index is None:
                return {"error": "Each operation requires 'index' (1-based)"}
            index = int(index)
            count = max(1, min(int(count), _MAX_ROWS))

            if op == "insert_rows":
                for _ in range(count):
                    ws.Rows(index).Insert()
            elif op == "delete_rows":
                rng = ws.Range(ws.Rows(index), ws.Rows(index + count - 1))
                rng.Delete()
            elif op == "insert_cols":
                for _ in range(count):
                    ws.Columns(index).Insert()
            elif op == "delete_cols":
                rng = ws.Range(ws.Columns(index), ws.Columns(index + count - 1))
                rng.Delete()
            ops_done += 1

        wb.Save()
    finally:
        wb.Close(SaveChanges=False)

    return {
        "result": f"Completed {ops_done} insert/delete operation(s) on {display_path}",
        "path": display_path,
        "operations_completed": ops_done,
    }


def _insert_delete_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    insert_delete_ops: list[dict[str, Any]] = kwargs.get("insert_delete_ops") or []
    if not insert_delete_ops:
        return {"error": "insert_delete_ops is required for insert_delete action"}
    if len(insert_delete_ops) > _MAX_EDIT_OPS:
        return {"error": f"Too many insert/delete operations (max {_MAX_EDIT_OPS})"}

    wb, err = _open_workbook_lib(resolved, display_path)
    if err:
        return {"error": err}

    ws, ws_err = _get_sheet_lib(wb, kwargs.get("sheet_name"))
    if ws_err:
        wb.close()
        return {"error": ws_err}

    ops_done = 0
    for op_def in insert_delete_ops:
        op = op_def.get("op", "")
        index = op_def.get("index")
        count = op_def.get("count", 1)
        if op not in _VALID_INSERT_DELETE_OPS:
            wb.close()
            return {"error": f"Unknown op: '{op}'. Valid: {', '.join(_VALID_INSERT_DELETE_OPS)}"}
        if index is None:
            wb.close()
            return {"error": "Each operation requires 'index' (1-based)"}
        index = int(index)
        count = max(1, min(int(count), _MAX_ROWS))

        if op == "insert_rows":
            ws.insert_rows(index, count)
        elif op == "delete_rows":
            ws.delete_rows(index, count)
        elif op == "insert_cols":
            ws.insert_cols(index, count)
        elif op == "delete_cols":
            ws.delete_cols(index, count)
        ops_done += 1

    wb.save(resolved)
    wb.close()
    return {
        "result": f"Completed {ops_done} insert/delete operation(s) on {display_path}",
        "path": display_path,
        "operations_completed": ops_done,
    }


# ---------------------------------------------------------------------------
# copy_range — copy/paste cell ranges
# ---------------------------------------------------------------------------


def _copy_range_com(manager: Any, resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    source_range = kwargs.get("source_range")
    dest_cell = kwargs.get("dest_cell")
    if not source_range or not dest_cell:
        return {"error": "source_range and dest_cell are required for copy_range action"}

    _, wb, err = _open_workbook_com(manager, resolved, display_path)
    if err:
        return {"error": err}

    try:
        ws_src, ws_err = _get_sheet_com(wb, kwargs.get("sheet_name"))
        if ws_err:
            return {"error": ws_err}

        dest_sheet = kwargs.get("dest_sheet")
        if dest_sheet:
            ws_dst, ws_err = _get_sheet_com(wb, dest_sheet)
            if ws_err:
                return {"error": ws_err}
        else:
            ws_dst = ws_src

        copy_values_only = kwargs.get("copy_values_only", False)
        src_rng = ws_src.Range(source_range)

        if copy_values_only:
            src_rng.Copy()
            ws_dst.Range(dest_cell).PasteSpecial(Paste=-4163)  # xlPasteValues
        else:
            src_rng.Copy(Destination=ws_dst.Range(dest_cell))

        wb.Save()
    finally:
        wb.Close(SaveChanges=False)

    return {
        "result": f"Copied {source_range} to {dest_cell} in {display_path}",
        "path": display_path,
        "source_range": source_range,
        "dest_cell": dest_cell,
        "values_only": copy_values_only,
    }


def _copy_range_lib(resolved: str, display_path: str, **kwargs: Any) -> dict[str, Any]:
    source_range = kwargs.get("source_range")
    dest_cell = kwargs.get("dest_cell")
    if not source_range or not dest_cell:
        return {"error": "source_range and dest_cell are required for copy_range action"}

    wb, err = _open_workbook_lib(resolved, display_path)
    if err:
        return {"error": err}

    ws_src, ws_err = _get_sheet_lib(wb, kwargs.get("sheet_name"))
    if ws_err:
        wb.close()
        return {"error": ws_err}

    dest_sheet = kwargs.get("dest_sheet")
    if dest_sheet:
        ws_dst, ws_err = _get_sheet_lib(wb, dest_sheet)
        if ws_err:
            wb.close()
            return {"error": ws_err}
    else:
        ws_dst = ws_src

    copy_values_only = kwargs.get("copy_values_only", False)

    try:
        from openpyxl.utils import coordinate_to_tuple

        src_cells = ws_src[source_range]
        if not isinstance(src_cells, tuple):
            src_cells = ((src_cells,),)

        dest_row, dest_col = coordinate_to_tuple(dest_cell)

        cells_copied = 0
        for r_offset, row in enumerate(src_cells):
            for c_offset, cell in enumerate(row):
                dst_cell = ws_dst.cell(row=dest_row + r_offset, column=dest_col + c_offset)
                if copy_values_only:
                    dst_cell.value = cell.value
                else:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        dst_cell.value = cell.value
                    else:
                        dst_cell.value = cell.value
                    if cell.has_style:
                        from copy import copy as _copy

                        dst_cell.font = _copy(cell.font)
                        dst_cell.fill = _copy(cell.fill)
                        dst_cell.border = _copy(cell.border)
                        dst_cell.alignment = _copy(cell.alignment)
                        dst_cell.number_format = cell.number_format
                        dst_cell.protection = _copy(cell.protection)
                cells_copied += 1
    except Exception as exc:
        wb.close()
        return {"error": f"Copy failed: {type(exc).__name__}: {exc}"}

    wb.save(resolved)
    wb.close()
    return {
        "result": f"Copied {source_range} to {dest_cell} in {display_path}",
        "path": display_path,
        "source_range": source_range,
        "dest_cell": dest_cell,
        "values_only": copy_values_only,
        "cells_copied": cells_copied,
    }


# ---------------------------------------------------------------------------
# Validation helpers
# ---------------------------------------------------------------------------


_ALLOWED_URL_SCHEMES = ("http://", "https://", "mailto:")

_DDE_PATTERN = re.compile(r"^[=+\-@]\s*[A-Za-z]+\|", re.IGNORECASE)


def _sanitize_cell_value(val: Any) -> Any:
    """Block DDE formula injection in COM cell writes.

    Values like ``=CMD|'/C calc'!A0`` exploit DDE to execute system commands.
    Legitimate formulas (``=SUM(A1:A10)``) do not contain pipe characters after
    the function name, so this only blocks the DDE vector.
    """
    if isinstance(val, str) and _DDE_PATTERN.match(val):
        return "'" + val
    return val


def _validate_url(url: str) -> str | None:
    """Return an error string if *url* uses a disallowed scheme, else None."""
    if not any(url.lower().startswith(s) for s in _ALLOWED_URL_SCHEMES):
        return f"Invalid URL scheme. Only http://, https://, and mailto: are allowed. Got: {url}"
    return None


# ---------------------------------------------------------------------------
# Color helpers
# ---------------------------------------------------------------------------


def _parse_color_int(color: str) -> int:
    """Parse hex color string to COM RGB integer (BGR format)."""
    c = color.lstrip("#")
    if len(c) == 6:
        r, g, b = int(c[0:2], 16), int(c[2:4], 16), int(c[4:6], 16)
        return r + (g << 8) + (b << 16)
    return 0


def _normalize_hex_color(color: str) -> str:
    """Normalize color to 6-char hex without #."""
    c = color.lstrip("#")
    if len(c) == 3:
        c = "".join(ch * 2 for ch in c)
    return c.upper()
